import os
import datetime

from dotenv import load_dotenv

# Load environment variables
load_dotenv()

import streamlit as st
from dotenv import load_dotenv
from openai import OpenAI
from PIL import Image

from src.ai_content_detector_ui import ai_content_detector_ui
from src.chat_ui import text_based
from src.cucp_reevals import cucp_reevaluations
from src.foundation_model_chat import foundation_model_chat_ui
from src.highway_incident_summarizer import summarize_caltrans_incidents
from src.landing_ai_ui import render_landing_ai_evaluation_ui
from src.project_delivery_evaluator import (
    extract_text_from_uploaded_pdf as pde_extract_pdf,
    extract_multi_doc_context,
    load_delivery_method_kb,
    run_delivery_evaluation,
    compute_delivery_recommendation,
    build_evaluation_excel,
    build_evaluation_excel_v2,
    score_all_methods,
    run_validation_analysis,
    ALL_METHODS,
    RUBRIC_QUESTIONS,
)
from src.pde_pipeline import run_pde_evaluation


def are_all_selected(options_list, selected_fields):
    return all(option in selected_fields for option in options_list)


# Set up the page layout
st.set_page_config(page_title="Caltrans", layout="wide")

# Apply blue sidebar styling
st.markdown(
    """
    <style>
    [data-testid="stSidebar"] {
        background-color: #3b82c8;
    }
    </style>
""",
    unsafe_allow_html=True,
)

# Load and apply custom CSS for general styling
with open("style/final.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# Display company logo (centered)
_, logo_col, _ = st.columns([1, 2, 1])
with logo_col:
    st.image("image/caltrans.jpg", use_container_width=True)

# Header
st.markdown(
    "<p style='text-align: center; color: black; margin-top: -10px; font-size: 30px;'>"
    "<span style='font-weight: bold'>CUCP: Transforming Documents into Insights with Agentic AI</span></p>",
    unsafe_allow_html=True,
)

# Horizontal line
st.markdown(
    "<hr style='height: 2.5px; margin-top: 0px; width: 100%; background-color: gray; margin-left: auto; margin-right: auto;'>",
    unsafe_allow_html=True,
)

# Sidebar
with st.sidebar:
    st.markdown(
        "<p style='text-align: center; color: white; font-size:25px;'><span style='font-weight: bold; font-family: century-gothic';></span>Solutions Scope</p>",
        unsafe_allow_html=True,
    )
    vAR_AI_application = st.selectbox(
        "Application Selection", ["Select Application", "Caltrans"], key="application", label_visibility="collapsed"
    )
    vAR_LLM_model = st.selectbox(
        "LLM Model Selection",
        [
            "LLM Models",
            "gpt-3.5-turbo-16k-0613",
            "gpt-4-0314",
            "gpt-3.5-turbo-16k",
            "gpt-3.5-turbo-1106",
            "gpt-4-0613",
            "gpt-4-0314",
        ],
        key="text_llmmodel",
        label_visibility="collapsed"
    )
    vAR_LLM_framework = st.selectbox(
        "LLM Framework Selection", ["LLM Framework", "Langchain"], key="text_framework", label_visibility="collapsed"
    )
    vAR_Gcp_cloud = st.selectbox(
        "GCP Services Selection",
        ["GCP Services Used", "VM Instance", "Computer Engine", "Cloud Storage"],
        key="text2",
        label_visibility="collapsed"
    )
    st.markdown("#### ")
    href = """<form action="#">
        <input type="submit" value="Clear/Reset"/>
        </form>"""
    st.sidebar.markdown(href, unsafe_allow_html=True)
    st.markdown("# ")
    st.markdown(
        "<p style='text-align: center; color: White; font-size:20px;'>Build & Deployed on<span style='font-weight: bold'></span></p>",
        unsafe_allow_html=True,
    )
    s1, s2, s3 = st.columns((4, 4, 4))
    with s1:
        st.image("image/oie_png.png")
    with s2:
        st.image("image/aws_logo.png")
    with s3:
        st.image("image/AzureCloud_img.png")


# Layouts
col1, col2, col3, col4, col5 = st.columns((2, 5, 2, 5, 2))
col21, col22, col23, col24, col25 = st.columns((2, 5, 2, 5, 2))
col61, col62, col63, col64, col65 = st.columns((2, 5, 5, 5, 2))
col71, col72, col73 = st.columns([1, 7, 1])

# Only show usecase selection if application is selected
if vAR_AI_application == "Caltrans":
    with col2:
        st.subheader("Select Application")
        st.write("#")

    with col4:
        app_option = st.selectbox(
            "Usecase Selection",
            (
                "Select the Usecase",
                "AI Content Detector",
                "CUCP Re-Evaluations",
                "Foundation Model",
                "Guardrails",
                "Highway Incident Summarizer Bot",
                "Human in the feedback Loop",
                "Langchain",
                "LLM as a Judge",
                "LLM Training",
                "LLM Evaluation",
                "Prompt Engineering",
                "RAG-Document Intelligence",
                "Personal Narrative Insights",
                "Project Delivery Evaluator",
                "Project Delivery Evaluator V2",
                "Right of Way (ROW) Evaluation",
            ),
            key="app_select",
            label_visibility="collapsed"
        )
        st.write("## ")
else:
    # Show instruction to select application first
    with col2:
        st.subheader("Select Application")
        st.write("#")
    with col4:
        st.info("Please select the application from the sidebar to continue")
        app_option = "Select the Usecase"  # Default value

# Main Routing
if app_option != "Select the Usecase":
    # Handle external links - Auto-open in new tab
    if app_option == "AI Content Detector":
        ai_content_detector_ui()

    elif app_option == "Langchain":
        st.markdown(
            """
            <script>
                window.open('https://promptwithlangchain-398219119144.us-east1.run.app/', '_blank');
            </script>
        """,
            unsafe_allow_html=True,
        )

        st.info("🔗 Opening Langchain application in a new tab...")
        st.markdown(
            """
            <div style="margin-top: 15px;">
                <p style="font-size: 14px; color: #666; margin-bottom: 8px;">
                    If the new tab didn't open automatically:
                </p>
                <a href="https://promptwithlangchain-398219119144.us-east1.run.app/" target="_blank" rel="noopener noreferrer" style="text-decoration: none;">
                    <button style="
                        background: linear-gradient(135deg, #3b82c8 0%, #2563eb 100%);
                        color: white;
                        padding: 8px 16px;
                        border: none;
                        border-radius: 6px;
                        cursor: pointer;
                        font-size: 14px;
                        font-weight: 500;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                        transition: all 0.3s ease;
                    " onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 4px 8px rgba(0,0,0,0.15)';"
                    onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 4px rgba(0,0,0,0.1)';">
                        Open Langchain App ↗
                    </button>
                </a>
            </div>
        """,
            unsafe_allow_html=True,
        )
        st.stop()

    elif app_option == "Prompt Engineering":
        st.markdown(
            """
            <script>
                window.open('https://genaiandpromptingtechnic-398219119144.us-east1.run.app/', '_blank');
            </script>
        """,
            unsafe_allow_html=True,
        )

        st.info("🔗 Opening Prompt Engineering application in a new tab...")
        st.markdown(
            """
            <div style="margin-top: 15px;">
                <p style="font-size: 14px; color: #666; margin-bottom: 8px;">
                    If the new tab didn't open automatically:
                </p>
                <a href="https://genaiandpromptingtechnic-398219119144.us-east1.run.app/" target="_blank" rel="noopener noreferrer" style="text-decoration: none;">
                    <button style="
                        background: linear-gradient(135deg, #3b82c8 0%, #2563eb 100%);
                        color: white;
                        padding: 8px 16px;
                        border: none;
                        border-radius: 6px;
                        cursor: pointer;
                        font-size: 14px;
                        font-weight: 500;
                        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                        transition: all 0.3s ease;
                    " onmouseover="this.style.transform='translateY(-2px)'; this.style.boxShadow='0 4px 8px rgba(0,0,0,0.15)';"
                    onmouseout="this.style.transform='translateY(0)'; this.style.boxShadow='0 2px 4px rgba(0,0,0,0.1)';">
                        Open Prompt Engineering App ↗
                    </button>
                </a>
            </div>
        """,
            unsafe_allow_html=True,
        )
        st.stop()

    elif app_option == "Guardrails":
        text_based("Guardrails", "None")

    elif app_option == "RAG-Document Intelligence":
        with col22:
            st.subheader("Upload the Policy Document")
        with col24:
            knowledge_base = st.file_uploader(
                "Upload Knowledge Base",
                type=["pdf", "txt"],
                key="knowledge_base_upload",
            )
        text_based("RAG-Document Intelligence", knowledge_base)

    elif app_option == "Human in the feedback Loop":
        with col22:
            st.subheader("Upload the Policy Document")
        with col24:
            knowledge_base = st.file_uploader(
                "Upload Knowledge Base",
                type=["pdf", "txt"],
                key="knowledge_base_upload",
            )
            
        st.markdown("<br>", unsafe_allow_html=True)
        # Wrap the whole feedback component in a bordered container with rounded edges
        feedback_container = st.container(border=True)
        with feedback_container:
            st.markdown(
                "<h3 style='margin-bottom: 0;'>Human in the Feedback Loop</h3>"
                "<hr style='margin-top: 5px; margin-bottom: 20px;'>", 
                unsafe_allow_html=True
            )
            text_based("Human in the feedback Loop", knowledge_base)

    elif app_option == "LLM Training":
        with col22:
            st.subheader("Upload Training Data")
        with col24:
            training_file = st.file_uploader(
                "Upload Training Dataset (CSV, JSON, or JSONL)",
                type=["jsonl", "json", "csv"],
                key="training_file_upload",
            )
        text_based("LLM Training", training_file)

    elif app_option == "CUCP Re-Evaluations":
        with col22:
            st.subheader("Upload the Narrative(s)")
        with col24:
            cucp_files = st.file_uploader(
                "Upload Narrative(s)",
                type=["pdf"],
                key="cucp_upload",
                accept_multiple_files=True,
            )

        with col22:
            st.subheader("Upload PNW Data")
        with col24:
            revenue_excel = st.file_uploader(
                "Upload PNW Data",
                type=["xlsx", "xls"],
                key="revenue_upload",
                help="Excel file with a 'Firms' sheet containing columns: Firm Name and Five Year Average",
            )

        # Parse the Excel file if provided
        firm_revenues = {}
        if revenue_excel:
            import pandas as pd
            import re
            try:
                # Read all sheets without assuming where the header is
                all_sheets = pd.read_excel(revenue_excel, sheet_name=None, header=None)
                
                target_df = None
                firm_col_name = None
                avg_col_name = None
                
                # Look specifically for the sheet containing "Firms" in its name (case-insensitive)
                sheet_names = list(all_sheets.keys())
                firms_sheet_name = next((name for name in sheet_names if "firms" in str(name).lower()), None)
                
                # If we found it, check that one. Otherwise, fallback to checking all of them.
                sheets_to_check = [firms_sheet_name] if firms_sheet_name else sheet_names
                
                for sheet_name in sheets_to_check:
                    if not sheet_name:
                        continue
                        
                    df = all_sheets[sheet_name]
                    
                    # Scan the first 15 rows to find the exact row with our headers
                    for i in range(min(15, len(df))):
                        # Convert to lowercase and normalize spaces (removes \n, \t, and extra spaces)
                        row_values = [re.sub(r'\s+', ' ', str(val)).strip().lower() for val in df.iloc[i].values]
                        
                        # Fuzzy matching: check if "firm" and "name" are in a cell, and "5" or "five" and "average" are in another
                        has_firm = any("firm" in v and "name" in v for v in row_values)
                        has_avg = any(("five" in v or "5" in v) and "average" in v for v in row_values)
                        
                        if has_firm and has_avg:
                            # Found the header! Set the columns and grab the data below it
                            df.columns = df.iloc[i]
                            target_df = df.iloc[i+1:].reset_index(drop=True)
                            
                            # Clean up column names internally to strip messy Excel spacing
                            cleaned_cols = [re.sub(r'\s+', ' ', str(c)).strip() for c in target_df.columns]
                            target_df.columns = cleaned_cols
                            
                            # Identify the exact column names the file uses to avoid KeyErrors
                            firm_col_name = next((c for c in cleaned_cols if "firm" in c.lower() and "name" in c.lower()), None)
                            avg_col_name = next((c for c in cleaned_cols if ("five" in c.lower() or "5" in c.lower()) and "average" in c.lower()), None)
                            break
                            
                    if target_df is not None and firm_col_name and avg_col_name:
                        break  # Stop looking once we've found our data
                
                # Now process the found data
                if target_df is not None and firm_col_name and avg_col_name:
                    target_df = target_df.dropna(subset=[firm_col_name])
                    
                    for _, row in target_df.iterrows():
                        firm_name = str(row[firm_col_name]).strip()
                        raw_revenue = row[avg_col_name]
                        
                        if firm_name.lower() == 'nan' or not firm_name:
                            continue
                            
                        # Handle formatting (e.g., "$1,234.50" -> float)
                        if isinstance(raw_revenue, str):
                            clean_rev = raw_revenue.replace("$", "").replace(",", "").strip()
                            try:
                                rev_val = float(clean_rev)
                            except ValueError:
                                rev_val = None
                        else:
                            try:
                                rev_val = float(raw_revenue)
                            except (ValueError, TypeError):
                                rev_val = None
                                
                        if rev_val is not None:
                            firm_revenues[firm_name] = rev_val
                else:
                    st.warning("Could not find columns resembling 'Firm Name' and 'Five Year Average'. Please ensure they exist in the 'Firms' sheet.")
            
            except Exception as e:
                st.error(f"Error parsing Excel file: {e}")

        # --- Memory Manager Sidebar ---
        with st.sidebar:
            
            # Push the feedback section lower in the sidebar
            st.markdown("###")
            st.markdown("###")
            st.markdown("---")
            
            from src.memory_manager import get_precedent_count, consolidate_memory_via_llm, overwrite_db
            
            total_precedents = get_precedent_count(1) + get_precedent_count(2) + get_precedent_count(3)
            pct = int((total_precedents / 135) * 100) if total_precedents > 0 else 0
            
            # Auto-consolidate if any level has hit the 45 limit
            any_level_full = any(get_precedent_count(lvl) >= 45 for lvl in [1, 2, 3])
            if any_level_full and not st.session_state.get('_auto_consolidated'):
                with st.spinner("Auto-consolidating corrections (a level reached its limit)..."):
                    clean_json_str = consolidate_memory_via_llm()
                    st.session_state['consolidated_rules_json'] = clean_json_str
                    st.session_state['show_consolidation_success'] = True
                    st.session_state['_auto_consolidated'] = True
            
            # --- Feedback + Merge Card ---
            # Inject a stable marker right before the container
            st.markdown('<div id="feedback-card-anchor"></div>', unsafe_allow_html=True)
            
            # CSS to style the adjacent container wrapper cleanly via sibling selector
            st.markdown("""
            <style>
            div[data-testid="stElementContainer"]:has(#feedback-card-anchor) + div[data-testid="stElementContainer"] div[data-testid="stVerticalBlockBorderWrapper"] {
                border: 2px solid rgba(255,255,255,0.35) !important;
                border-radius: 14px !important;
                background: rgba(255,255,255,0.06) !important;
                padding: 18px 16px !important;
            }
            </style>
            """, unsafe_allow_html=True)
            
            # Create a native Streamlit border container so it safely wraps the button
            feedback_card = st.container(border=True)
            
            with feedback_card:
                # Feedback Header
                st.markdown("""
                <div style="background: rgba(255,255,255,0.12); padding: 12px 16px; border-radius: 10px; text-align: center; margin-bottom: 14px;">
                    <p style="margin: 0; color: white; font-size: 1rem; font-weight: 600; letter-spacing: 0.04em; text-transform: uppercase;">Human in the Feedback Loop</p>
                </div>
                <p style="color: rgba(255,255,255,0.75); font-size: 0.82rem; margin: 0 0 14px 0; line-height: 1.5;">The AI learns from your corrections. Your edits are remembered and applied to all future evaluations.</p>
                """, unsafe_allow_html=True)
                
                # Metrics Area
                st.markdown(f"""
                <div style="background: rgba(255,255,255,0.08); border-radius: 10px; padding: 14px 16px; margin-bottom: 18px;">
                    <div style="display: flex; justify-content: space-between; align-items: baseline; margin-bottom: 8px;">
                        <p style="margin: 0; color: rgba(255,255,255,0.7); font-size: 0.78rem; font-weight: 500; text-transform: uppercase;">Your Saved Corrections</p>
                        <p style="margin: 0; color: rgba(255,255,255,0.5); font-size: 0.72rem;">max 45 per step</p>
                    </div>
                    <p style="margin: 0 0 10px 0; color: white; font-size: 1.8rem; font-weight: 700;">{total_precedents}<span style="color: rgba(255,255,255,0.4); font-size: 1rem; font-weight: 400;"> / 135 total</span></p>
                    <div style="background: rgba(255,255,255,0.15); border-radius: 6px; height: 6px; overflow: hidden;">
                        <div style="background: linear-gradient(90deg, #60a5fa, #a78bfa); height: 100%; width: {pct}%; border-radius: 6px;"></div>
                    </div>
                    <p style="margin: 6px 0 0 0; color: rgba(255,255,255,0.5); font-size: 0.7rem; line-height: 1.4;">Each evaluation step can store up to 45 corrections. When full, corrections are auto-merged into a clean rulebook.</p>
                </div>
                """, unsafe_allow_html=True)
    
                # Separator
                st.markdown('<div style="height: 1px; background: rgba(255,255,255,0.12); margin: 0 0 14px 0;"></div>', unsafe_allow_html=True)
                
                # UX-5: Correction Timeline
                all_staged = (st.session_state.get("staged_precedents", {}).get("level_1_precedents", []) +
                              st.session_state.get("staged_precedents", {}).get("level_2_precedents", []) +
                              st.session_state.get("staged_precedents", {}).get("level_3_precedents", []))
                a_overrides = st.session_state.get('analyst_overrides', [])
                if all_staged or a_overrides:
                    with st.expander("🕒 Correction History (This Session)", expanded=False):
                        if a_overrides:
                            st.markdown("**Analyst Overrides:**")
                            for ao in a_overrides:
                                st.markdown(f"- {ao['field']}: {ao['value']}")
                        if all_staged:
                            st.markdown("**AI Precedents:**")
                            for p in all_staged:
                                target_display = p.get('target', 'Unknown')
                                st.markdown(f"- Overrode **{str(target_display)[:30]}** ➔ {p.get('correction')}")
                    st.markdown('<div style="height: 1px; background: rgba(255,255,255,0.12); margin: 14px 0;"></div>', unsafe_allow_html=True)
    
                # Merge Section
                st.markdown("""
                <div style="background: rgba(255,255,255,0.08); padding: 10px 14px; border-radius: 10px; text-align: center; margin-bottom: 8px;">
                    <p style="margin: 0; color: rgba(255,255,255,0.7); font-size: 0.88rem; font-weight: 600; text-transform: uppercase;">Merge Corrections</p>
                </div>
                <p style="color: rgba(255,255,255,0.55); font-size: 0.78rem; margin: 0 0 12px 0; line-height: 1.5;">Combine all corrections into a single clean rulebook file that you can download and re-upload later.</p>
                """, unsafe_allow_html=True)
                
                if total_precedents < 15:
                    st.markdown("""
                    <div style="
                        background: rgba(255,255,255,0.06);
                        padding: 10px 14px;
                        border-radius: 8px;
                        border: 1px solid rgba(255,255,255,0.1);
                        margin: 0 0 14px 0;
                    ">
                        <p style="margin: 0; color: rgba(255,255,255,0.65); font-size: 0.78rem; line-height: 1.5;">
                            Collect at least <strong style="color: rgba(255,255,255,0.9);">15 corrections</strong> before merging for best results.
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Button is now INSIDE the bordered card
                if st.button("Merge & Download Rulebook", use_container_width=True):
                    with st.spinner("AI is merging your corrections..."):
                        clean_json_str = consolidate_memory_via_llm()
                        st.session_state['consolidated_rules_json'] = clean_json_str
                        st.session_state['show_consolidation_success'] = True
                        
                if st.session_state.get('show_consolidation_success') and 'consolidated_rules_json' in st.session_state:
                    st.markdown("""
                    <div style="
                        background: rgba(74, 222, 128, 0.15);
                        padding: 10px 14px;
                        border-radius: 8px;
                        border: 1px solid rgba(74, 222, 128, 0.3);
                        margin: 10px 0;
                    ">
                        <p style="margin: 0; color: rgba(255,255,255,0.9); font-size: 0.82rem;">&#x2705; Merge complete. Download your rulebook below.</p>
                    </div>
                    """, unsafe_allow_html=True)
                    st.download_button(
                        label="Download Rulebook",
                        data=st.session_state['consolidated_rules_json'],
                        file_name=f"cucp_rules_{datetime.date.today().isoformat()}.json",
                        mime="application/json",
                        use_container_width=True
                    )

        if not cucp_files:
            st.markdown("""
            <div style="
                background: linear-gradient(135deg, #f0f7ff 0%, #e8f0fe 100%);
                border: 1px solid #bcd4f0;
                border-radius: 12px;
                padding: 24px 28px;
                margin: 20px 0;
            ">
                <h3 style="margin: 0 0 12px 0; color: #1a3d6e;">👋 Welcome! Here's how to get started:</h3>
                <ol style="margin: 0; padding-left: 20px; color: #333; line-height: 2;">
                    <li><strong>Upload the applicant's narrative</strong> using the uploader above</li>
                    <li><strong>Optionally upload PNW data</strong> for automatic cross-referencing</li>
                    <li>Click <strong>Start AI Evaluation</strong> — the AI will walk you through each step</li>
                </ol>
            </div>
            """, unsafe_allow_html=True)

        if cucp_files:
            if len(cucp_files) > 1:
                st.warning(f"⚠️ You uploaded {len(cucp_files)} files. Only the first file will be processed in interactive mode. Process one narrative at a time.")
            cucp_file = cucp_files[0]  # Interactive mode processes one document at a time
            file_name = cucp_file.name
            base_name = file_name.rsplit(".", 1)[0]
            
            # --- Detect File Change ---
            if 'current_file_name' not in st.session_state:
                st.session_state.current_file_name = file_name
            elif st.session_state.current_file_name != file_name:
                # File changed! Wipe staged precedents and process state
                for key in ['eval_stage', 'pdf_text', 'l1_data', 'l2_data', 'l3_data', 'staged_precedents', 'analyst_overrides']:
                    if key in st.session_state:
                        del st.session_state[key]
                st.session_state.current_file_name = file_name
                st.rerun()
            
            # --- Initialize State Machine ---
            if 'eval_stage' not in st.session_state:
                st.session_state.eval_stage = 0
            if 'staged_precedents' not in st.session_state:
                st.session_state.staged_precedents = {
                    "level_1_precedents": [],
                    "level_2_precedents": [],
                    "level_3_precedents": []
                }
            if 'analyst_overrides' not in st.session_state:
                st.session_state.analyst_overrides = []
            if 'pdf_text' not in st.session_state:
                from src.cucp_reevals import extract_text_from_pdf
                extracted = extract_text_from_pdf(cucp_file)
                if not extracted or not extracted.strip():
                    st.error("⚠️ Could not extract any text from this PDF. The file may be scanned/image-based or corrupted. Please upload a text-based PDF.")
                    st.stop()
                st.session_state.pdf_text = extracted
            
            from src.cucp_reevals import run_level_1_extraction, run_level_2_classification, run_level_3_thresholds, generate_final_md_report
            from src.memory_manager import add_precedent
            
            # --- UI Styling: borders, buttons ---
            st.markdown("""
            <style>
            /* Thicker border on correction/feedback expanders */
            details[data-testid="stExpander"] {
                border: 2px solid rgba(59,130,246,0.25) !important;
                border-radius: 12px !important;
            }
            /* Rounded border around evaluation container */
            div[data-testid="stVerticalBlock"]:has(> div.eval-border-marker) {
                border: 1.5px solid rgba(148,163,184,0.3);
                border-radius: 16px;
                padding: 24px 20px 20px 20px;
            }
            .eval-border-marker { display: none; }
            
            /* Modern button styling */
            div.stButton > button {
                border-radius: 10px !important;
                font-weight: 600 !important;
                padding: 0.5rem 1.2rem !important;
                transition: all 0.2s ease !important;
                border: 1.5px solid transparent !important;
            }
            /* Primary buttons (Approve, Start) */
            div.stButton > button[kind="primary"] {
                background: #3b82f6 !important;
                color: white !important;
                border-color: #3b82f6 !important;
            }
            div.stButton > button[kind="primary"]:hover {
                background: #2563eb !important;
                border-color: #2563eb !important;
                box-shadow: 0 4px 14px rgba(59,130,246,0.35) !important;
            }
            /* Secondary buttons (Go Back, etc.) */
            div.stButton > button[kind="secondary"] {
                background: #f1f5f9 !important;
                color: #475569 !important;
                border-color: #e2e8f0 !important;
            }
            div.stButton > button[kind="secondary"]:hover {
                background: #e2e8f0 !important;
                box-shadow: 0 2px 8px rgba(0,0,0,0.08) !important;
            }
            </style>
            """, unsafe_allow_html=True)
            
            # --- Progress Stepper ---
            st.markdown('<div class="eval-border-marker"></div>', unsafe_allow_html=True)
            with st.container(border=True):
                st.markdown("<h4 style='text-align: center; margin-bottom: 0px; color: #1e293b; font-weight: 600;'>Human in the Feedback Loop</h4>", unsafe_allow_html=True)
                step_labels = ["Start", "Step 1: Review Facts", "Step 2: Review Classifications", "Step 3: Review Thresholds", "Final Report"]
                current_step = st.session_state.eval_stage
                stepper_html = '<div style="display: flex; align-items: center; justify-content: center; margin: 10px 0 10px 0; gap: 0;">'
                for i, label in enumerate(step_labels):
                    if i < current_step:
                        color = "#22c55e"; bg = "rgba(34,197,94,0.15)"; border_c = "#22c55e"; icon = "✓"
                    elif i == current_step:
                        color = "#3b82f6"; bg = "rgba(59,130,246,0.15)"; border_c = "#3b82f6"; icon = str(i)
                    else:
                        color = "#94a3b8"; bg = "rgba(148,163,184,0.08)"; border_c = "#cbd5e1"; icon = str(i)
                    stepper_html += f'<div style="display:flex; flex-direction:column; align-items:center; min-width:80px;">'
                    stepper_html += f'<div style="width:36px; height:36px; border-radius:50%; background:{bg}; border:2px solid {border_c}; display:flex; align-items:center; justify-content:center; font-size:0.85rem; font-weight:700; color:{color};">{icon}</div>'
                    stepper_html += f'<p style="margin:4px 0 0 0; font-size:0.82rem; color:{color}; text-align:center; line-height:1.3;">{label}</p></div>'
                    if i < len(step_labels) - 1:
                        line_color = "#22c55e" if i < current_step else "#cbd5e1"
                        stepper_html += f'<div style="flex:1; height:3px; background:{line_color}; margin:0 4px; margin-bottom:22px;"></div>'
                stepper_html += '</div>'
                st.markdown(stepper_html, unsafe_allow_html=True)
            
            # STATE 0: Upload & Process Level 1
            if st.session_state.eval_stage == 0:
                st.markdown("### Getting Started")
                with st.expander("📁 Have a saved rulebook from a previous session? Upload it here.", expanded=False):
                    uploaded_rules = st.file_uploader("Upload your previously saved corrections rulebook (optional)", type=['json'])
                    if uploaded_rules is not None:
                        try:
                            import json
                            rules_dict = json.load(uploaded_rules)
                            overwrite_db(rules_dict)
                            st.success("Successfully loaded your previous corrections!")
                        except Exception as e:
                            st.error(f"Failed to load rulebook: {e}")
                        
                st.markdown("---")
                if st.button("Start AI Evaluation ➔", type="primary"):
                    with st.spinner(f"Running Step 1: Reading and extracting facts from **{file_name}**..."):
                        l1_result = run_level_1_extraction(
                            st.session_state.pdf_text, 
                            firm_revenues, 
                            staged_precedents=st.session_state.staged_precedents.get("level_1_precedents", [])
                        )
                    if "error" in l1_result:
                        st.error(f"⚠️ Something went wrong during fact extraction: {l1_result['error']}")
                    else:
                        st.session_state.l1_data = l1_result
                        # Save original values for undo support
                        st.session_state._original_firm_name = l1_result.get('firm_name', 'None')
                        st.session_state._original_narrative_pnw = l1_result.get('narrative_pnw', 'NOT PROVIDED')
                        st.session_state.eval_stage = 1
                        st.rerun()
            
            # STATE 1: Level 1 Fact Review
            elif st.session_state.eval_stage == 1:
                st.markdown("### Step 1: Review Extracted Facts")
                st.info("The AI has read the narrative and pulled out the key facts. Please check if any facts are wrong or missing. If everything looks correct, click **Approve & Continue** at the bottom.")
                st.caption("*Tip: Double-click any cell in the table below to expand it and read the full text.*")
                
                # Display Data
                import pandas as pd
                l1_data = st.session_state.l1_data
                
                # Show Firm Info & Financials separately before the table
                st.markdown(f"**Found Applicant Firm:** `{l1_data.get('firm_name', 'None')}`")
                st.markdown(f"**Excel Cross-Reference Revenue/PNW:** `{l1_data.get('cross_reference_result', 'None')}`")
                st.markdown(f"**Narrative Declared PNW:** `{l1_data.get('narrative_pnw', 'NOT PROVIDED')}`")
                st.markdown("---")
                
                facts = l1_data.get('extracted_facts', [])
                df_facts = pd.DataFrame(facts)
                # Human-readable column headers
                col_rename = {
                    'id': 'Fact #', 'when': 'When', 'where': 'Where',
                    'who': 'Who', 'what': 'What', 'why': 'Why',
                    'magnitude': 'Magnitude/Threshold', 'demographic_flag': 'Demographics Checkbox',
                    'source_quote': 'Source Quote'
                }
                df_facts = df_facts.rename(columns={k: v for k, v in col_rename.items() if k in df_facts.columns})
                st.dataframe(df_facts, use_container_width=True)

                st.markdown("---")
                
                l1_count = get_precedent_count(1) + len(st.session_state.staged_precedents.get("level_1_precedents", []))
                
                if st.button("⬅️ Go Back to Start"):
                    st.session_state.eval_stage = 0
                    st.rerun()
                
                # Correction Form (collapsed by default)
                with st.expander("✏️ Structural issue? Click here to correct via AI re-run", expanded=False):
                    
                    # Undo + Clear buttons at top of expander
                    if l1_count > 0 or len(st.session_state.get('analyst_overrides', [])) > 0:
                        uc1, uc2 = st.columns(2)
                        with uc1:
                            if st.button("↩️ Undo Last Correction", key="l1_undo"):
                                if st.session_state.staged_precedents.get("level_1_precedents"):
                                    st.session_state.staged_precedents["level_1_precedents"].pop()
                                    with st.spinner("Re-running Fact Extraction..."):
                                        l1_result = run_level_1_extraction(
                                            st.session_state.pdf_text, firm_revenues, staged_precedents=st.session_state.staged_precedents["level_1_precedents"]
                                        )
                                        st.session_state.l1_data = l1_result
                                    st.rerun()
                                elif st.session_state.get('analyst_overrides'):
                                    undone = st.session_state.analyst_overrides.pop()
                                    if undone.get('field') == 'Firm Name':
                                        st.session_state.l1_data['firm_name'] = st.session_state.get('_original_firm_name', 'None')
                                    elif undone.get('field') == 'Narrative Declared PNW':
                                        st.session_state.l1_data['narrative_pnw'] = st.session_state.get('_original_narrative_pnw', 'NOT PROVIDED')
                                    st.rerun()
                        with uc2:
                            if st.button("🗑️ Clear All Step Corrections", key="l1_clear"):
                                st.session_state.staged_precedents["level_1_precedents"] = []
                                st.session_state.analyst_overrides = []
                                st.session_state.l1_data['firm_name'] = st.session_state.get('_original_firm_name', 'None')
                                st.session_state.l1_data['narrative_pnw'] = st.session_state.get('_original_narrative_pnw', 'NOT PROVIDED')
                                with st.spinner("Re-running Fact Extraction from scratch..."):
                                    l1_result = run_level_1_extraction(
                                        st.session_state.pdf_text, firm_revenues, staged_precedents=[]
                                    )
                                    st.session_state.l1_data = l1_result
                                st.rerun()
                    
                    if l1_count >= 45:
                        st.error(f"🚨 Correction limit reached ({l1_count}/45). Your corrections will be auto-merged. Check the sidebar to download the rulebook.")
                    elif l1_count >= 36:
                        st.warning(f"⚠️ Approaching correction limit ({l1_count}/45). Corrections will be auto-merged when the limit is reached.")
                    
                    # Target selectors OUTSIDE form so they trigger reruns
                    c1, c2 = st.columns(2)
                    with c1:
                        fact_labels = [f"Fact {f.get('id', i+1)}" for i, f in enumerate(facts)]
                        target_type_options = fact_labels + ["Firm Name", "Narrative Declared PNW", "Specific Incident Detail"]
                        target_type = st.selectbox(
                            "What to Correct", 
                            target_type_options,
                            help="Select which fact or metadata field needs correction."
                        )
                    with c2:
                        if target_type.startswith("Fact "):
                            field_options = ["When", "Where", "Who", "What", "Why", "Magnitude"]
                        else:
                            field_options = ["Not Applicable"]
                        target_field = st.selectbox(
                            "Which Field",
                            field_options,
                            help="Select the specific field within this fact to correct." if target_type.startswith("Fact ") else "No field selection needed for this target.",
                            disabled=not target_type.startswith("Fact ")
                        )
                    
                    if target_type.startswith("Fact "):
                        target_fact = f"{target_type}: {target_field}"
                    else:
                        target_fact = target_type
                    
                    # Context-sensitive label
                    if target_fact == "Firm Name":
                        val_label = "Correct Firm Name"
                    elif target_fact == "Narrative Declared PNW":
                        val_label = "Correct PNW Amount"
                    elif target_fact == "Specific Incident Detail":
                        val_label = "Describe the missing or incorrect incident"
                    else:
                        val_label = "Corrected Value"
                    
                    with st.form("l1_correction_form", clear_on_submit=True):
                        correction_val = st.text_input(
                            val_label,
                            help="Enter the exact fact that the AI should have extracted."
                        )
                        reasoning_val = st.text_area(
                                "Reasoning for Change (What should the AI remember?)",
                                help="Explain why this is wrong. The AI will remember this and apply it to all future evaluations."
                            )
                        
                        # Disable submit if over threshold
                        is_disabled = True if l1_count >= 45 else False
                        if st.form_submit_button("Apply Correction & Re-Evaluate", disabled=is_disabled):
                            if not correction_val or not correction_val.strip():
                                st.error("Please enter a corrected value before submitting.")
                            elif not reasoning_val or not reasoning_val.strip():
                                st.error("Please provide reasoning so the AI can learn from your correction.")
                            elif target_fact in ["Narrative Declared PNW", "Firm Name"]:
                                field_key = "narrative_pnw" if target_fact == "Narrative Declared PNW" else "firm_name"
                                st.session_state.l1_data[field_key] = correction_val
                                st.session_state.analyst_overrides.append({
                                    "field": target_fact,
                                    "value": correction_val,
                                    "reasoning": reasoning_val
                                })
                                st.rerun()
                            else:
                                st.session_state.staged_precedents["level_1_precedents"].append({
                                    "target": target_fact,
                                    "correction": correction_val,
                                    "human_reasoning": reasoning_val
                                })
                                # Auto re-run Level 1 with the new correction and stay on review
                                with st.spinner(f"Re-running Fact Extraction with your correction..."):
                                    l1_result = run_level_1_extraction(
                                        st.session_state.pdf_text,
                                        firm_revenues,
                                        staged_precedents=st.session_state.staged_precedents.get("level_1_precedents", [])
                                    )
                                    st.session_state.l1_data = l1_result
                                st.rerun()
                
                st.caption("*Your corrections are saved only after you approve the final evaluation at the end.*")
                
                # Proceed Button
                excel_pnw = st.session_state.l1_data.get("cross_reference_result", "None")
                narrative_pnw = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                combined_financials = f"Excel Cross-Reference Revenue/PNW: {excel_pnw}\nNarrative Declared PNW: {narrative_pnw}"
                
                if st.button("Approve & Continue ➔", type="primary"):
                    with st.spinner("Running Step 2: Legal Classification..."):
                        l2_result = run_level_2_classification(
                            st.session_state.l1_data.get('extracted_facts', []),
                            combined_financials,
                            staged_precedents=st.session_state.staged_precedents.get("level_2_precedents", [])
                        )
                    if "error" in l2_result:
                        st.error(f"⚠️ Something went wrong during classification: {l2_result['error']}")
                    else:
                        st.session_state.l2_data = l2_result
                        st.session_state.eval_stage = 2
                        st.rerun()

            # STATE 2: Level 2 Classification Review
            elif st.session_state.eval_stage == 2:
                st.markdown("### Step 2: Review Legal Classifications")
                st.info("The AI has categorized each fact under [49 CFR §26.67](https://www.ecfr.gov/current/title-49/subtitle-A/part-26/subpart-D/section-26.67) (e.g., Social Disadvantage, Economic Disadvantage, Institutional/Systemic Barrier). Review the categories below. If a fact is in the wrong category, click the correction section to fix it. Otherwise, click **Approve & Continue**.")
                st.caption("*Tip: Double-click any cell in the table below to expand it and read the full text.*")
                
                l2_data = st.session_state.l2_data
                classifications = l2_data.get('classifications', [])
                df_class = pd.DataFrame(classifications)
                # Human-readable column headers
                l2_col_rename = {
                    'fact_id': 'Fact #', 'classification': 'Legal Category',
                    'summary': 'Summary', 'reasoning': 'AI Reasoning'
                }
                df_class = df_class.rename(columns={k: v for k, v in l2_col_rename.items() if k in df_class.columns})
                st.dataframe(df_class, use_container_width=True)
                
                l2_count = get_precedent_count(2) + len(st.session_state.staged_precedents.get("level_2_precedents", []))
                
                if st.button("⬅️ Go Back to Step 1"):
                    st.session_state.eval_stage = 1
                    st.rerun()
                
                with st.expander("✏️ Wrong category? Click here to reclassify a fact", expanded=False):
                    
                    # Undo + Clear buttons at top of expander
                    if l2_count > 0:
                        uc1, uc2 = st.columns(2)
                        with uc1:
                            if st.button("↩️ Undo Last Correction", key="l2_undo"):
                                if st.session_state.staged_precedents.get("level_2_precedents"):
                                    st.session_state.staged_precedents["level_2_precedents"].pop()
                                    with st.spinner("Re-evaluating Classifications..."):
                                        excel_pnw_rerun = st.session_state.l1_data.get("cross_reference_result", "None")
                                        narrative_pnw_rerun = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                                        combined_financials_rerun = f"Excel Cross-Reference Revenue/PNW: {excel_pnw_rerun}\nNarrative Declared PNW: {narrative_pnw_rerun}"
                                        st.session_state.l2_data = run_level_2_classification(
                                            st.session_state.l1_data.get('extracted_facts', []),
                                            combined_financials_rerun,
                                            staged_precedents=st.session_state.staged_precedents.get("level_2_precedents", [])
                                        )
                                    st.rerun()
                        with uc2:
                            if st.button("🗑️ Clear All Step Corrections", key="l2_clear"):
                                st.session_state.staged_precedents["level_2_precedents"] = []
                                with st.spinner("Re-evaluating Classifications from scratch..."):
                                    excel_pnw_rerun = st.session_state.l1_data.get("cross_reference_result", "None")
                                    narrative_pnw_rerun = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                                    combined_financials_rerun = f"Excel Cross-Reference Revenue/PNW: {excel_pnw_rerun}\nNarrative Declared PNW: {narrative_pnw_rerun}"
                                    st.session_state.l2_data = run_level_2_classification(
                                        st.session_state.l1_data.get('extracted_facts', []),
                                        combined_financials_rerun,
                                        staged_precedents=[]
                                    )
                                st.rerun()
                    
                    if l2_count >= 45:
                        st.error(f"🚨 Correction limit reached ({l2_count}/45). Your corrections will be auto-merged. Check the sidebar to download the rulebook.")
                    elif l2_count >= 36:
                        st.warning(f"⚠️ Approaching correction limit ({l2_count}/45). Corrections will be auto-merged when the limit is reached.")
                    
                    # Target selectors OUTSIDE form so they trigger reruns
                    c1, c2 = st.columns(2)
                    with c1:
                        target_options = [f"Fact {c.get('fact_id', i+1)} — {c.get('classification', c.get('Legal Category', ''))}" for i, c in enumerate(classifications)]
                        target_options += ["General Scenario Misclassification"]
                        target_class = st.selectbox(
                            "Fact to Correct", 
                            target_options,
                            help="Select the specific AI classification you want to override."
                        )
                    with c2:
                        if target_class != "General Scenario Misclassification":
                            category_options = ["Keep Current (Fix Reasoning Only)", "Social Disadvantage", "Economic Disadvantage", "Institutional/Systemic Barrier", "Ordinary Business Risk", "Insufficient Evidence"]
                        else:
                            category_options = ["Not Applicable — describe in reasoning below"]
                        correction_class = st.selectbox(
                            "New Category", 
                            category_options,
                            help="Select 'Keep Current' if the category is correct but the reasoning needs fixing. Otherwise pick the correct category." if target_class != "General Scenario Misclassification" else "For general guidance, describe the pattern in the reasoning field below.",
                            disabled=target_class == "General Scenario Misclassification"
                        )
                    
                    with st.form("l2_correction_form", clear_on_submit=True):
                        reasoning_class = st.text_area(
                                "Legal Rationale (Why is this the correct classification?)",
                                help="Provide your legal justification. The AI will learn this and apply it to similar cases in the future."
                            )
                        
                        is_disabled = True if l2_count >= 45 else False
                        if st.form_submit_button("Apply Override & Re-Evaluate", disabled=is_disabled):
                            if not reasoning_class or not reasoning_class.strip():
                                st.error("Please provide a legal rationale so the AI can learn from your correction.")
                            else:
                                # Extract fact_id from the target string for structured storage
                                selected_idx = target_options.index(target_class) if target_class in target_options else None
                                resolved_fact_id = classifications[selected_idx].get('fact_id', selected_idx + 1) if selected_idx is not None and selected_idx < len(classifications) else None
                                # Resolve "Keep Current" to the actual current classification
                                final_correction = correction_class
                                if correction_class == "Keep Current (Fix Reasoning Only)" and selected_idx is not None and selected_idx < len(classifications):
                                    final_correction = classifications[selected_idx].get('classification', correction_class)
                                st.session_state.staged_precedents["level_2_precedents"].append({
                                    "target": target_class,
                                    "fact_id": resolved_fact_id,
                                    "correction": final_correction,
                                    "human_reasoning": reasoning_class
                                })
                                # Re-run with FULL financial context preserved
                                excel_pnw_rerun = st.session_state.l1_data.get("cross_reference_result", "None")
                                narrative_pnw_rerun = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                                combined_financials_rerun = f"Excel Cross-Reference Revenue/PNW: {excel_pnw_rerun}\nNarrative Declared PNW: {narrative_pnw_rerun}"
                                with st.spinner("Re-evaluating Classifications..."):
                                    st.session_state.l2_data = run_level_2_classification(
                                        st.session_state.l1_data.get('extracted_facts', []),
                                        combined_financials_rerun,
                                        staged_precedents=st.session_state.staged_precedents.get("level_2_precedents", [])
                                    )
                                st.rerun()
                            
                st.caption("*Your corrections are saved only after you approve the final evaluation at the end.*")

                excel_pnw = st.session_state.l1_data.get("cross_reference_result", "None")
                narrative_pnw = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                combined_financials = f"Excel Cross-Reference Revenue/PNW: {excel_pnw}\nNarrative Declared PNW: {narrative_pnw}"

                if st.button("Approve & Continue ➔", type="primary"):
                    with st.spinner("Running Step 3: Evidentiary Thresholds..."):
                        l3_result = run_level_3_thresholds(
                            st.session_state.l2_data.get("classifications", []), 
                            st.session_state.l1_data.get("extracted_facts", []), 
                            combined_financials,
                            staged_precedents=st.session_state.staged_precedents.get("level_3_precedents", [])
                        )
                    if "error" in l3_result:
                        st.error(f"⚠️ Something went wrong during threshold evaluation: {l3_result['error']}")
                    else:
                        st.session_state.l3_data = l3_result
                        st.session_state.eval_stage = 3
                        st.rerun()

            # STATE 3: Level 3 Threshold Review
            elif st.session_state.eval_stage == 3:
                st.markdown("### Step 3: Review Pass/Fail Decisions")
                st.info("The AI has evaluated the 7 mandatory CUCP criteria under [49 CFR §26.67](https://www.ecfr.gov/current/title-49/subtitle-A/part-26/subpart-D/section-26.67) and assigned Pass or Fail to each. The standard of proof is **preponderance of evidence** (more likely true than not, >50%). Review the decisions below. If the AI is too strict or too lenient on any criterion, click the correction section. Otherwise, click **Approve & Finalize**.")
                st.caption("*Tip: Double-click any cell in the table below to expand it and read the full text.*")
                
                l3_data = st.session_state.l3_data
                criteria = l3_data.get('criteria', [])
                df_crit = pd.DataFrame(criteria)
                
                # Human-readable column headers
                l3_col_rename = {
                    's_no': '#', 'category': 'Category', 'qualification': 'Criterion',
                    'evidence_summary': 'Evidence Summary', 'reasoning': 'AI Reasoning',
                    'pass_fail': 'Pass/Fail', 'request_info': 'Need More Info?',
                    'confidence': 'Confidence'
                }
                df_crit = df_crit.rename(columns={k: v for k, v in l3_col_rename.items() if k in df_crit.columns})
                
                # Change 5: Round confidence to 1 decimal (safe for non-numeric)
                if 'Confidence' in df_crit.columns:
                    def _safe_round(x):
                        try:
                            return round(float(x), 1)
                        except (ValueError, TypeError):
                            return x
                    df_crit['Confidence'] = df_crit['Confidence'].apply(_safe_round)
                
                cols_to_show = ['Category', 'Criterion', 'Evidence Summary', 'AI Reasoning', 'Pass/Fail', 'Need More Info?', 'Confidence']
                existing_cols = [c for c in cols_to_show if c in df_crit.columns]
                
                # UX-2: Font-only visual highlighting (minimal, clean)
                def _styler_pass_fail(val):
                    if str(val).lower() == 'pass': return 'color: #22c55e; font-weight: bold;'
                    elif str(val).lower() == 'fail': return 'color: #ef4444; font-weight: bold;'
                    return ''
                def _styler_confidence(val):
                    try:
                        v = float(val)
                        if v >= 8.0: return 'color: #22c55e; font-weight: bold;'
                        elif v >= 5.0: return 'color: #f59e0b; font-weight: bold;'
                        else: return 'color: #ef4444; font-weight: bold;'
                    except: return ''
                    
                if existing_cols and 'Pass/Fail' in existing_cols and 'Confidence' in existing_cols:
                    styled_df = df_crit[existing_cols].style.map(_styler_pass_fail, subset=['Pass/Fail']).map(_styler_confidence, subset=['Confidence'])
                    st.dataframe(styled_df, use_container_width=True)
                elif existing_cols:
                    st.dataframe(df_crit[existing_cols], use_container_width=True)
                else:
                    st.warning("The AI returned no evaluation criteria. Please go back and try again.")
                
                l3_count = get_precedent_count(3) + len(st.session_state.staged_precedents.get("level_3_precedents", []))
                
                if st.button("⬅️ Go Back to Step 2"):
                    st.session_state.eval_stage = 2
                    st.rerun()
                
                st.write(f"**Final Evaluated Decision:** {l3_data.get('final_decision')}")
                with st.expander("✏️ Disagree with a decision? Click here to adjust", expanded=False):
                    
                    # Undo + Clear buttons at top of expander
                    if l3_count > 0:
                        uc1, uc2 = st.columns(2)
                        with uc1:
                            if st.button("↩️ Undo Last Correction", key="l3_undo"):
                                if st.session_state.staged_precedents.get("level_3_precedents"):
                                    st.session_state.staged_precedents["level_3_precedents"].pop()
                                    with st.spinner("Re-evaluating decisions..."):
                                        excel_pnw_rerun = st.session_state.l1_data.get("cross_reference_result", "None")
                                        narrative_pnw_rerun = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                                        combined_financials_rerun = f"Excel Cross-Reference Revenue/PNW: {excel_pnw_rerun}\nNarrative Declared PNW: {narrative_pnw_rerun}"
                                        st.session_state.l3_data = run_level_3_thresholds(
                                            st.session_state.l2_data.get("classifications", []), 
                                            st.session_state.l1_data.get("extracted_facts", []), 
                                            combined_financials_rerun,
                                            staged_precedents=st.session_state.staged_precedents.get("level_3_precedents", [])
                                        )
                                    st.rerun()
                        with uc2:
                            if st.button("🗑️ Clear All Step Corrections", key="l3_clear"):
                                st.session_state.staged_precedents["level_3_precedents"] = []
                                with st.spinner("Re-evaluating decisions from scratch..."):
                                    excel_pnw_rerun = st.session_state.l1_data.get("cross_reference_result", "None")
                                    narrative_pnw_rerun = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                                    combined_financials_rerun = f"Excel Cross-Reference Revenue/PNW: {excel_pnw_rerun}\nNarrative Declared PNW: {narrative_pnw_rerun}"
                                    st.session_state.l3_data = run_level_3_thresholds(
                                        st.session_state.l2_data.get("classifications", []), 
                                        st.session_state.l1_data.get("extracted_facts", []), 
                                        combined_financials_rerun,
                                        staged_precedents=[]
                                    )
                                st.rerun()
                    
                    if l3_count >= 45:
                        st.error(f"🚨 Correction limit reached ({l3_count}/45). Your corrections will be auto-merged. Check the sidebar to download the rulebook.")
                    elif l3_count >= 36:
                        st.warning(f"⚠️ Approaching correction limit ({l3_count}/45). Corrections will be auto-merged when the limit is reached.")
                    
                    # Target selectors OUTSIDE form so they trigger reruns
                    target_crit = st.selectbox(
                        "Criterion or Decision to Adjust", 
                        [f"{c.get('category', c.get('Category', ''))} — {c.get('qualification', c.get('Criterion', ''))}" for c in criteria] + ["Final Decision"],
                        help="Select the specific criterion where the AI was too strict or too lenient. Each criterion corresponds to a mandatory eligibility requirement under 49 CFR §26.67 for Social and Economic Disadvantage (SED) determinations."
                    )
                    if target_crit == "Final Decision":
                        result_options = ["Pass", "Fail"]
                    else:
                        result_options = ["Keep Current (Fix Reasoning Only)", "Pass", "Fail", "Request Additional Information"]
                    correction_crit = st.selectbox(
                        "New Result", 
                        result_options,
                        help="For Final Decision, only Pass or Fail applies. For individual criteria, select 'Keep Current' to fix only the reasoning." if target_crit == "Final Decision" else "Select 'Keep Current' if the decision is correct but the reasoning needs fixing."
                    )
                    
                    with st.form("l3_correction_form", clear_on_submit=True):
                        reasoning_crit = st.text_area(
                            "Why is this the correct decision?",
                            help="Explain your reasoning. The AI will remember this standard and apply it consistently in future reviews."
                        )
                        
                        is_disabled = True if l3_count >= 45 else False
                        if st.form_submit_button("Apply Correction & Re-Evaluate", disabled=is_disabled):
                            if not reasoning_crit or not reasoning_crit.strip():
                                st.error("Please explain your reasoning so the AI can learn from your correction.")
                            else:
                                # Resolve "Keep Current" to the actual current decision
                                final_correction = correction_crit
                                if correction_crit == "Keep Current (Fix Reasoning Only)":
                                    crit_options = [f"{c.get('category', c.get('Category', ''))} — {c.get('qualification', c.get('Criterion', ''))}" for c in criteria]
                                    if target_crit in crit_options:
                                        selected_crit_idx = crit_options.index(target_crit)
                                        final_correction = criteria[selected_crit_idx].get('pass_fail', correction_crit)
                                st.session_state.staged_precedents["level_3_precedents"].append({
                                    "target": target_crit,
                                    "correction": final_correction,
                                    "human_reasoning": reasoning_crit
                                })
                                # Re-run with FULL combined financial context (narrative PNW override preserved)
                                excel_pnw_rerun = st.session_state.l1_data.get("cross_reference_result", "None")
                                narrative_pnw_rerun = st.session_state.l1_data.get("narrative_pnw", "NOT PROVIDED")
                                combined_financials_rerun = f"Excel Cross-Reference Revenue/PNW: {excel_pnw_rerun}\nNarrative Declared PNW: {narrative_pnw_rerun}"
                                with st.spinner("Re-evaluating decisions..."):
                                    st.session_state.l3_data = run_level_3_thresholds(
                                        st.session_state.l2_data.get("classifications", []), 
                                        st.session_state.l1_data.get("extracted_facts", []), 
                                        combined_financials_rerun,
                                        staged_precedents=st.session_state.staged_precedents.get("level_3_precedents", [])
                                    )
                                st.rerun()
                            
                st.caption("*Your corrections are saved only after you approve the final evaluation below.*")

                if st.button("Approve Final Evaluation & Commit Corrections ➔", type="primary"):
                    from src.memory_manager import commit_staged_precedents
                    commit_staged_precedents(st.session_state.staged_precedents)
                    st.session_state.eval_stage = 4
                    st.rerun()
                    
            # STATE 4: Final Generation
            elif st.session_state.eval_stage == 4:
                st.success(f"✅ Full Process-Supervised Evaluation completed for **{file_name}**")
                
                result_md = generate_final_md_report(
                    st.session_state.l1_data, 
                    st.session_state.l2_data, 
                    st.session_state.l3_data, 
                    st.session_state.get('analyst_overrides', [])
                )
                
                # Display Results
                st.markdown("---")
                st.markdown(result_md, unsafe_allow_html=True)
                
                colA, colB = st.columns(2)
                with colA:
                    if st.button("⬅️ Go Back to Level 3"):
                        st.session_state.eval_stage = 3
                        st.rerun()
                with colB:
                    if st.button("🔄 Reset / Start Over"):
                        for key in ['eval_stage', 'pdf_text', 'l1_data', 'l2_data', 'l3_data', 'staged_precedents', 'analyst_overrides', 'consolidated_rules_json', 'show_consolidation_success', 'current_file_name', '_auto_consolidated', '_original_firm_name', '_original_narrative_pnw']:
                            if key in st.session_state:
                                del st.session_state[key]
                        st.rerun()

                # --- Handle Excel Export (Old Logic adapted for new State) ---
                from io import BytesIO
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                import re

                def parse_md_tables(md_text):
                    tables = []
                    lines = md_text.split("\n")
                    i = 0
                    while i < len(lines):
                        line = lines[i].strip()
                        if "|" in line and re.search(r"[a-zA-Z]", line):
                            header = [c.strip() for c in line.strip("|").split("|")]
                            if i + 1 < len(lines) and re.match(r"[\|\s\-:]+$", lines[i+1].strip()):
                                rows = []
                                j = i + 2
                                while j < len(lines):
                                    row_line = lines[j].strip()
                                    if "|" not in row_line or not row_line:
                                        break
                                    row = [c.strip() for c in row_line.strip("|").split("|")]
                                    rows.append(row)
                                    j += 1
                                if rows:
                                    tables.append((header, rows))
                                i = j
                                continue
                        i += 1
                    return tables

                def build_excel(md_text, report_title):
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)
                    tables = parse_md_tables(md_text)
                    sheet_names = ["Part 1 – Evaluation", "Part 2 – Explainable AI"]
                    header_fill = PatternFill("solid", fgColor="1F4E79")
                    header_font = Font(bold=True, color="FFFFFF", size=11)
                    border_side = Side(style="thin", color="000000")
                    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
                    even_fill = PatternFill("solid", fgColor="EBF3FB")

                    for t_idx, (headers, rows) in enumerate(tables):
                        sheet_name = sheet_names[t_idx] if t_idx < len(sheet_names) else f"Table {t_idx+1}"
                        ws = wb.create_sheet(title=sheet_name)
                        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
                        title_cell = ws.cell(row=1, column=1, value=report_title)
                        title_cell.font = Font(bold=True, size=13, color="1F4E79")
                        title_cell.alignment = Alignment(horizontal="center", vertical="center")
                        ws.row_dimensions[1].height = 22
                        for col_i, h in enumerate(headers, start=1):
                            cell = ws.cell(row=2, column=col_i, value=h)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                            cell.border = cell_border
                        ws.row_dimensions[2].height = 20
                        for row_i, row in enumerate(rows, start=3):
                            fill = even_fill if row_i % 2 == 0 else PatternFill()
                            for col_i, val in enumerate(row, start=1):
                                cell = ws.cell(row=row_i, column=col_i, value=val)
                                cell.border = cell_border
                                cell.alignment = Alignment(wrap_text=True, vertical="top")
                                if fill.fill_type:
                                    cell.fill = fill
                        for col_i, h in enumerate(headers, start=1):
                            col_letter = openpyxl.utils.get_column_letter(col_i)
                            max_len = max(len(h), *[len(str(row[col_i - 1])) if col_i - 1 < len(row) else 0 for row in rows])
                            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

                    comments_match = re.search(r'### 📝 CERTIFIER COMMENTS & FINAL SUMMARY\s*(.*?)(?=\Z)', md_text, re.DOTALL | re.IGNORECASE)
                    if comments_match:
                        comments_text = comments_match.group(1).strip()
                        if comments_text:
                            ws_comments = wb.create_sheet(title="Final Summary")
                            ws_comments.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
                            title_cell = ws_comments.cell(row=1, column=1, value="Certifier Comments & Final Summary")
                            title_cell.font = Font(bold=True, size=13, color="1F4E79")
                            title_cell.alignment = Alignment(horizontal="center", vertical="center")
                            ws_comments.row_dimensions[1].height = 22
                            content_cell = ws_comments.cell(row=3, column=1, value=comments_text)
                            content_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                            ws_comments.column_dimensions['A'].width = 100
                            approx_lines = (len(comments_text) // 100) + comments_text.count("\n") + 2
                            ws_comments.row_dimensions[3].height = max(15, approx_lines * 15)

                    # Change 12: Extract Analyst Overrides & Classification Summary bullet lists
                    overrides_match = re.search(r'### 🧑‍⚖️ ANALYST OVERRIDES\s*(.*?)(?=###|---|\Z)', md_text, re.DOTALL)
                    class_match = re.search(r'### 🏷️ CLASSIFICATION SUMMARY\s*(.*?)(?=###|---|\Z)', md_text, re.DOTALL)
                    if overrides_match or class_match:
                        ws_extra = wb.create_sheet(title="Overrides & Classifications")
                        ws_extra.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
                        title_cell = ws_extra.cell(row=1, column=1, value="Analyst Overrides & Classification Summary")
                        title_cell.font = Font(bold=True, size=13, color="1F4E79")
                        title_cell.alignment = Alignment(horizontal="center", vertical="center")
                        ws_extra.row_dimensions[1].height = 22
                        current_row = 3
                        if overrides_match:
                            ws_extra.cell(row=current_row, column=1, value="Analyst Overrides").font = Font(bold=True, size=11, color="1F4E79")
                            current_row += 1
                            for line in overrides_match.group(1).strip().split("\n"):
                                line = line.strip().lstrip("- ")
                                if line:
                                    ws_extra.cell(row=current_row, column=1, value=re.sub(r'\*+', '', line))
                                    ws_extra.row_dimensions[current_row].height = 18
                                    current_row += 1
                            current_row += 1
                        if class_match:
                            ws_extra.cell(row=current_row, column=1, value="Classification Summary").font = Font(bold=True, size=11, color="1F4E79")
                            current_row += 1
                            for line in class_match.group(1).strip().split("\n"):
                                line = line.strip().lstrip("- ")
                                if line:
                                    ws_extra.cell(row=current_row, column=1, value=re.sub(r'\*+', '', line))
                                    ws_extra.row_dimensions[current_row].height = 18
                                    current_row += 1
                        ws_extra.column_dimensions['A'].width = 80

                    buf = BytesIO()
                    wb.save(buf)
                    buf.seek(0)
                    return buf

                excel_buf = build_excel(result_md, f"CUCP Evaluation — {file_name}")
                dl_col1, dl_col2 = st.columns([1, 4])
                with dl_col1:
                    st.download_button(
                        label="📊 Download (.xlsx)",
                        data=excel_buf.getvalue(),
                        file_name=f"{base_name}_evaluation.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_xlsx_0",
                    )

    elif app_option == "LLM Evaluation":
        with col22:
            st.subheader("Upload the Policy Document")
        with col24:
            knowledge_base = st.file_uploader(
                "Upload Knowledge Base",
                type=["pdf", "txt"],
                key="knowledge_base_upload",
            )
        text_based("LLM Evaluation", knowledge_base)

    elif app_option == "Foundation Model":
        foundation_model_chat_ui()
    elif app_option == "Highway Incident Summarizer Bot":
        text_based("Highway Incident Summarizer Bot", None)

    elif app_option == "LLM as a Judge":
        with col22:
            st.subheader("Upload the Policy Document")
        with col24:
            policy_file = st.file_uploader(
                "Upload Knowledge Base",
                type=["pdf", "txt"],
                key="policy_file_upload",
            )
        text_based("LLM as a Judge", policy_file)

    elif app_option == "Personal Narrative Insights":
        with col22:
            st.subheader("Upload the Personal Narrative")
        with col24:
            knowledge_base = st.file_uploader(
                "Upload Personal Narrative",
                type=["pdf", "txt"],
                key="personal_narrative_upload",
            )
        text_based("Personal Narrative Insights", None)

    elif app_option == "Highway Incident Summarizer":
        highway_incident_ui(app_option)

    elif app_option == "Right of Way (ROW) Evaluation":
        with col22:
            st.subheader("Upload ROW Appraisal Report")
        with col24:
            st.file_uploader(
                "Upload Right of Way Appraisal PDF",
                type=["pdf"],
                key="row_upload",
                help="Upload a Caltrans Right of Way appraisal report for rubric-based evaluation.",
            )
        render_landing_ai_evaluation_ui()

    elif app_option in ["Project Delivery Evaluator", "Project Delivery Evaluator V2"]:
        is_pde_v2_menu = app_option == "Project Delivery Evaluator V2"
        with col22:
            st.subheader("Upload Nomination Fact Sheet")
        with col24:
            delivery_files = st.file_uploader(
                "Upload project nomination fact sheet(s) (.docx or .pdf)",
                type=["docx", "pdf"],
                accept_multiple_files=True,
                key="delivery_upload",
            )
            # Rulebook upload (optional — for institutional memory)
            st.markdown("<p style='font-size:0.85rem; margin-bottom:2px; color:#475569;'>Upload prior Rulebook (pde_rules.json) — optional</p>", unsafe_allow_html=True)
            rulebook_file = st.file_uploader(
                "Upload prior Rulebook",
                type=["json"],
                key="delivery_rulebook_upload",
                help="Upload a previously downloaded rulebook to apply institutional memory to this evaluation.",
                label_visibility="collapsed",
            )
            if rulebook_file is not None:
                from src.pde_memory_manager import load_rulebook
                _loaded_rules, _load_warn = load_rulebook(rulebook_file)
                if _load_warn and "Could not" in _load_warn:
                    st.error(f"Rulebook load error: {_load_warn}")
                elif _loaded_rules:
                    st.session_state.pde_rules = _loaded_rules
                    _approved_count = sum(1 for r in _loaded_rules if r.get("status") == "approved")
                    st.success(f"{_approved_count} approved rule(s) loaded — these will guide this evaluation.")
                    if _load_warn:
                        st.warning(f"Some rules were skipped: {_load_warn}")
            if "pde_rules" not in st.session_state:
                st.session_state.pde_rules = []

        with col71:
            st.write("")
        with col72:
            if not delivery_files:
                st.markdown("""
                <div style="
                    background: #ffffff;
                    border: 1.5px solid #bcd4f0;
                    border-radius: 14px;
                    padding: 28px 32px;
                    margin: 20px 0;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
                ">
                    <h3 style="margin: 0 0 14px 0; color: #1F4E79;">Project Delivery Method Evaluator</h3>
                    <p style="color: #334155; margin: 0 0 16px 0;">
                        This tool evaluates a project nomination fact sheet and recommends the most
                        appropriate delivery method based on the project's scope, schedule, cost,
                        innovation potential, and staffing considerations.
                    </p>
                    <ol style="color: #475569; margin: 0; padding-left: 20px; line-height: 1.8;">
                        <li><strong>Upload</strong> your completed nomination fact sheet (and any supporting docs).</li>
                        <li><strong>Evaluate</strong> to analyze the project characteristics.</li>
                        <li><strong>Review</strong> the ratings, evidence, and recommendation.</li>
                        <li><strong>Download</strong> the results as an Excel report.</li>
                    </ol>
                    <p style="color: #64748b; margin: 16px 0 0 0; font-size: 0.85rem;">
                        <strong>Delivery Methods:</strong> Design-Bid-Build, Design-Sequencing, Design-Build/Low-Bid,
                        Design-Build/Best-Value, CM/GC, Progressive Design-Build
                    </p>
                </div>
                """, unsafe_allow_html=True)
            else:
                # Extract text based on file type
                file_names = sorted([f.name for f in delivery_files])
                if "pde_current_files" not in st.session_state or st.session_state.pde_current_files != file_names:
                    # New files uploaded — reset state
                    for key in [k for k in list(st.session_state.keys()) if k.startswith("pde_")]:
                        del st.session_state[key]
                    st.session_state.pde_current_files = file_names

                    narrative_text = extract_multi_doc_context(delivery_files)
                    
                    # We no longer extract static ratings from docx tables.
                    # We only extract the text corpus.
                    narrative_text = extract_multi_doc_context(delivery_files)

                    st.session_state.pde_narrative = narrative_text
                    
                    # Initialize an empty structure for manual UI overrides
                    st.session_state.pde_manual_ratings = {}
                    st.session_state.pde_existing_ratings = {}

                narrative_text = st.session_state.get("pde_narrative", "")
                existing_ratings = st.session_state.get("pde_existing_ratings", {})

                # Show file loaded confirmation
                st.success(f"Loaded {len(delivery_files)} document(s): **{', '.join(file_names)}**")

                # Load knowledge base (cached)
                if "pde_kb_text" not in st.session_state:
                    with st.spinner("Preparing evaluation..."):
                        st.session_state.pde_kb_text = load_delivery_method_kb()
                
                # Role Selection — locked once evaluation is run
                _role_locked = "pde_eval_result" in st.session_state
                pde_role = st.radio(
                    "View Perspective",
                    ["District", "Headquarters (HQ)"],
                    horizontal=True,
                    key="pde_role",
                    disabled=_role_locked,
                )
                is_district_view = (pde_role == "District")
                pde_report_mode = "Template Summary + Method Sheets (V2)" if is_pde_v2_menu else "Current Report (V1)"

                # Run evaluation button
                if "pde_eval_result" not in st.session_state:
                    if st.button("Evaluate Project", type="primary", key="pde_run"):
                        with st.spinner("Analyzing project details..."):
                            eval_result = run_pde_evaluation(
                                narrative_text,
                                st.session_state.pde_kb_text,
                                existing_ratings if existing_ratings else None,
                                pde_rules=st.session_state.get("pde_rules", []),
                            )
                        if "error" in eval_result:
                            st.error(f"Evaluation Failed: {eval_result['error']}")
                        else:
                            st.session_state.pde_eval_result = eval_result
                            st.session_state.pde_recommendation = compute_delivery_recommendation(
                                eval_result.get("ratings", [])
                            )
                            st.session_state.pde_step = 1  # Start wizard at Step 1
                            st.rerun()

                # Display results
                if "pde_eval_result" in st.session_state:
                    import pandas as pd
                    eval_result = st.session_state.pde_eval_result
                    recommendation = st.session_state.pde_recommendation
                    ratings = eval_result.get("ratings", [])
                    missing = eval_result.get("missing_questions", [])
                    project_name = eval_result.get("project_name", file_names[0].rsplit(".", 1)[0] if file_names else "project")

                    # Compute multi-method scores (cached)
                    if "pde_multi_method" not in st.session_state:
                        st.session_state.pde_multi_method = score_all_methods(ratings)
                    multi_method_data = st.session_state.pde_multi_method

                    # --- Build override-merged ratings for dynamic re-scoring ---
                    _manual = st.session_state.get("pde_manual_ratings", {})
                    _has_overrides = bool(_manual)
                    if _has_overrides:
                        _merged_ratings = []
                        for _r in ratings:
                            _qid = _r.get("question_id", "")
                            if _qid in _manual:
                                _copy = dict(_r)
                                _copy["selected_rating"] = _manual[_qid]
                                _merged_ratings.append(_copy)
                            else:
                                _merged_ratings.append(_r)
                    else:
                        _merged_ratings = ratings

                    # Cache override-adjusted scoring — only recompute when manual ratings change
                    # This avoids firing an LLM API call on every Streamlit re-render.
                    _manual_key = str(sorted(_manual.items())) if _manual else ""
                    if (
                        "pde_override_multi" not in st.session_state
                        or st.session_state.get("pde_override_multi_key") != _manual_key
                    ):
                        if _has_overrides:
                            _override_rec = compute_delivery_recommendation(_merged_ratings)
                            _override_multi = score_all_methods(_merged_ratings)
                        else:
                            _override_rec = recommendation
                            _override_multi = multi_method_data
                        st.session_state.pde_override_rec = _override_rec
                        st.session_state.pde_override_multi = _override_multi
                        st.session_state.pde_override_multi_key = _manual_key
                    else:
                        _override_rec = st.session_state.pde_override_rec
                        _override_multi = st.session_state.pde_override_multi

                    # Compute validation analysis if user ratings exist (cached)
                    validation_data = None
                    if existing_ratings:
                        if "pde_validation" not in st.session_state:
                            st.session_state.pde_validation = run_validation_analysis(ratings, existing_ratings)
                        validation_data = st.session_state.pde_validation

                    # --- Recommendation Card (HQ only) ---
                    _ai_method = recommendation.get("recommended_method", "N/A")
                    _ai_score = recommendation.get("composite_score", 0)
                    _ai_runner = recommendation.get("runner_up_method", "N/A")
                    _ai_runner_aff = recommendation.get("runner_up_score")  # integer points, may be None
                    _ov_method = _override_rec.get("recommended_method", "N/A")
                    _ov_score = _override_rec.get("composite_score", 0)
                    _ov_runner = _override_rec.get("runner_up_method", "N/A")
                    _ov_runner_aff = _override_rec.get("runner_up_score")  # integer points, may be None
                    _has_overrides = bool(_manual)
                    _method_changed = _ai_method != _ov_method

                    # Active recommendation to display (override-adjusted if overrides exist)
                    rec_method = _ov_method
                    comp_score = _ov_score
                    runner_up = _ov_runner
                    runner_up_aff = _ov_runner_aff

                    if is_district_view:
                        # District recommendation card — no points, no scores
                        st.markdown(
                            f'<div style="background:#ffffff; border:1.5px solid #bcd4f0; border-radius:14px; padding:28px 36px; margin:10px 0 24px 0; box-shadow:0 2px 8px rgba(0,0,0,0.06);">'
                            f'<p style="color:#64748b; margin:0 0 8px 0; font-size:0.8rem; text-transform:uppercase; letter-spacing:1.5px; font-weight:600;">Recommended Delivery Method</p>'
                            f'<p style="color:#1F4E79; margin:0; font-size:2rem; font-weight:700; line-height:1.2;">{rec_method}</p>'
                            f'<p style="color:#64748b; margin:12px 0 0 0; font-size:0.95rem;">Alternative: <strong>{runner_up}</strong></p>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )
                    else:
                        # HQ recommendation card — full details with points
                        _runner_up_badge = ""
                        if runner_up_aff is not None:
                            _runner_up_badge = f'&nbsp;<span style="background:#f0f9ff; color:#0369a1; border:1px solid #bae6fd; border-radius:4px; padding:1px 6px; font-size:0.82rem; font-weight:600;">Score: {int(runner_up_aff)} pts</span>'
                        _adj_label = "  (Override-Adjusted)" if _has_overrides else ""

                        st.markdown(
                            f'<div style="background:#ffffff; border:1.5px solid #bcd4f0; border-radius:14px; padding:28px 36px; margin:10px 0 24px 0; box-shadow:0 2px 8px rgba(0,0,0,0.06);">'
                            f'<p style="color:#64748b; margin:0 0 8px 0; font-size:0.8rem; text-transform:uppercase; letter-spacing:1.5px; font-weight:600;">Recommended Delivery Method{_adj_label}</p>'
                            f'<div style="display:flex; justify-content:space-between; align-items:baseline; flex-wrap:wrap; gap:8px;">'
                            f'<p style="color:#1F4E79; margin:0; font-size:2rem; font-weight:700; line-height:1.2;">{rec_method}</p>'
                            f'<p style="color:#475569; margin:0; font-size:1.1rem; font-weight:500;">Score: {int(comp_score)} pts</p>'
                            f'</div>'
                            f'<p style="color:#64748b; margin:12px 0 0 0; font-size:0.95rem;">Alternative: <strong>{runner_up}</strong>{_runner_up_badge}</p>'
                            f'</div>',
                            unsafe_allow_html=True,
                        )

                        # Override impact badge (separate render to avoid nested HTML issues)
                        if _has_overrides and _method_changed:
                            st.markdown(
                                f'<div style="margin-top:-16px; margin-bottom:16px; padding:10px 14px; background:#fff7ed; border-left:4px solid #ea580c; border-radius:6px;">'
                                f'<p style="margin:0; font-size:0.8rem; color:#9a3412; font-weight:600;">Override Impact — Recommendation Changed</p>'
                                f'<p style="margin:4px 0 0 0; font-size:0.9rem; color:#374151;">AI Original: <strong>{_ai_method}</strong> ({int(_ai_score)} pts) &rarr; Override-Adjusted: <strong>{_ov_method}</strong> ({int(_ov_score)} pts)</p>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )
                        elif _has_overrides:
                            st.markdown(
                                f'<div style="margin-top:-16px; margin-bottom:16px; padding:10px 14px; background:#f0fdf4; border-left:4px solid #16a34a; border-radius:6px;">'
                                f'<p style="margin:0; font-size:0.8rem; color:#166534; font-weight:600;">Override Impact — Recommendation Confirmed</p>'
                                f'<p style="margin:4px 0 0 0; font-size:0.9rem; color:#374151;">Your {len(_manual)} override(s) were applied — the recommendation stays <strong>{_ov_method}</strong> (score adjusted from {int(_ai_score)} to {int(_ov_score)} pts).</p>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )


                    if pde_report_mode == "Template Summary + Method Sheets (V2)":
                        st.subheader("Template Questionnaire View")

                        # Use override-merged ratings so worksheet reflects user corrections
                        rating_index = {r.get("question_id"): r for r in _merged_ratings}
                        method_labels = [
                            "Design-Bid-Build",
                            "Design-Sequencing",
                            "Design-Build/Low-Bid",
                            "Design-Build/Best-Value",
                            "CM/GC",
                            "Progressive Design-Build",
                        ]
                        # === V2 TEMPLATE UI ===
                        st.markdown("### Project Delivery Selection Tool")
                        st.markdown("#### Project Summary Worksheet")

                        # 1. Instructions
                        st.markdown("""
                        <div style="border: 1px solid #ccc; padding: 10px; font-size: 0.85rem; background-color: #fcfcfc; margin-bottom: 20px;">
                        <strong>INSTRUCTIONS</strong><br/>
                        1. On the Project Summary Worksheet, complete the date of the review, project name, and selection committee members.<br/>
                        2. Answer all questions on Worksheet 1. Record the score for each delivery method on the form as indicated.<br/>
                        <i>Note: if any one of the answers is "No-Go," the delivery method need not be considered further for that project.</i><br/>
                        3. After all the questions are answered, total the score for each delivery system and transfer the totals to the Scoring Summary section on the Project Summary Worksheet.<br/>
                        4. Repeat steps 2 and 3 for Worksheet 2.<br/>
                        5. Total the scores from Worksheets 1 and 2 the in Scoring Summary section of the Project Summary Worksheet.<br/>
                        6. Select the project delivery method with the highest score and record any important selection committee comments in the space provided.<br/>
                        <br/>
                        <i>Note: Complete one project delivery selection questionnaire for each unique project...</i>
                        </div>
                        """, unsafe_allow_html=True)

                        # 2. Project Info
                        col1, col2 = st.columns([1, 2])
                        with col1:
                            st.text_input("Project Name", value=project_name, disabled=True)
                            st.text_input("Date of Review", value=datetime.date.today().strftime("%m/%d/%y"), disabled=True)
                        with col2:
                            st.text_area("Selection CommitteeMembers", value="", height=100, placeholder="Enter names...")
                        
                        st.write("---")

                        # 2. Project Headers
                        st.markdown(f"""
                        <div style="text-align: center; margin-bottom: 20px;">
                            <h2 style="margin: 0; font-size: 1.4rem;">Project Delivery Selection Tool</h2>
                            <h3 style="margin: 0; font-size: 1.2rem;">Project Summary Worksheet</h3>
                        </div>
                        <div style="margin-bottom: 15px; font-family: monospace;">
                            <div style="border-bottom: 1px solid #000; margin-bottom: 10px;">Project Name: {project_name}</div>
                            <div style="border-bottom: 1px solid #000; margin-bottom: 10px;">Date of Review: {datetime.date.today().strftime("%m/%d/%y")}</div>
                            <div style="border-bottom: 1px solid #000; height: 100px;">Selection Committee:</div>
                        </div>
                        """, unsafe_allow_html=True)
                        

                        # 4. Questionnaire
                        st.markdown("<hr style='border: 1px solid #000;'>", unsafe_allow_html=True)

                        from src.project_delivery_evaluator import get_selection_matrix_points

                        q_table_html = """
                        <style>
                            .parity-table { width: 100%; border-collapse: collapse; font-family: sans-serif; font-size: 0.85rem; border: 1px solid #000; }
                            .parity-table td, .parity-table th { border: 1px solid #000; padding: 8px; }
                            .parity-header { background-color: #f8f9fa; }
                            .id-cell { text-align: right; font-size: 0.7rem; padding-right: 2px !important; width: 40px; border-right: none !important; }
                            .box-cell { text-align: center; font-weight: bold; width: 55px; border-left: none !important; white-space: nowrap; }
                            .section-header { color: blue; font-weight: bold; background-color: #fcfcfc; }
                            .q-row-text { font-weight: bold; font-size: 0.9rem; width: 40%; }
                            .opt-row-text { font-size: 0.82rem; padding-left: 15px !important; width: 40%; }
                        </style>
                        <table class="parity-table">
                        """

                        # Add sub-header for methods (1st column + 6*2 method columns = 13 columns)
                        q_table_html += "<tr class='parity-header'><th></th>"
                        for m in method_labels:
                            q_table_html += f"<th colspan='2' style='text-align: center; font-weight: bold; font-size: 0.75rem;'>{m}</th>"
                        q_table_html += "</tr>"

                        # Worksheet 1
                        q_table_html += "<tr><td colspan='13' style='text-align: center; font-weight: bold; border-top: 2px solid #000; border-bottom: 2px solid #000;'>WORKSHEET 1 - PROJECT SCOPE AND CHARACTERISTICS</td></tr>"
                        q_table_html += "<tr><td colspan='13' class='section-header' style='background-color: #fff;'>A. Project Scope and Characteristic Criteria</td></tr>"

                        for q in RUBRIC_QUESTIONS:
                            qid = q["id"]
                            if qid == "B1": # Start Worksheet 2
                                q_table_html += "<tr><td colspan='13' style='height: 20px; border: none;'></td></tr>"
                                q_table_html += "<tr><td colspan='13' style='text-align: center; font-weight: bold; border: 1px solid #000;'>WORKSHEET 2 - EVALUATION OF SUCCESS CRITERIA</td></tr>"

                            if qid in ["B1", "C1", "D1", "E1", "F1"]:
                                section_titles = {
                                    "B1": "B. Schedule Issues",
                                    "C1": "C. Opportunity for Innovation",
                                    "D1": "D. Quality Enhancement",
                                    "E1": "E. Cost Issues",
                                    "F1": "F. Staffing Issues"
                                }
                                q_table_html += f"<tr><td colspan='13' class='section-header' style='background-color: #fff;'>{section_titles[qid]}</td></tr>"

                            robj = rating_index.get(qid, {})
                            sel_rating = robj.get("selected_rating", "").upper()

                            # Calculate rowspan (1 for question + number of options)
                            opt_keys = [o for o in ["option_a", "option_b", "option_c"] if q.get(o)]
                            rowspan = 1 + len(opt_keys)

                            # Question row
                            q_table_html += f"<tr><td class='q-row-text'>{qid}. {q['question']}</td>"
                            for m in method_labels:
                                if sel_rating in ("N/E", "NE", "N_E"):
                                    rating_display = "N/E"
                                elif not is_district_view:
                                    _pts = get_selection_matrix_points(qid, m, sel_rating, robj)
                                    if _pts is None:
                                        rating_display = "No-Go"
                                    else:
                                        rating_display = f"{sel_rating} ({_pts})" if sel_rating else ""
                                else:
                                    rating_display = sel_rating if sel_rating else ""
                                q_table_html += f"<td rowspan='{rowspan}' class='id-cell' style='vertical-align: top;'></td><td rowspan='{rowspan}' class='box-cell' style='vertical-align: top;'>{rating_display}</td>"
                            q_table_html += "</tr>"

                            # Option rows (Now only need the first column)
                            for opt_key in ["A", "B", "C"]:
                                opt_label = f"option_{opt_key.lower()}"
                                opt_text = q.get(opt_label, "")
                                if not opt_text: continue
                                q_table_html += f"<tr><td class='opt-row-text'>☐ {opt_key}. {opt_text}</td></tr>"


                        st.markdown(q_table_html, unsafe_allow_html=True)
                        
                        


                    # --- Override Notes and Close Match (HQ only — these reveal impact calculations) ---
                    if not is_district_view:
                        override_reasons = recommendation.get("override_reasons", [])
                        if override_reasons:
                            reasons_html = "".join(f"<li style='margin-bottom: 4px;'>{r}</li>" for r in override_reasons)
                            st.markdown(
                                f'<div style="background:#eff6ff; border-left:4px solid #3b82f6; border-radius:0 8px 8px 0; padding:16px 20px; margin:0 0 16px 0;">'
                                f'<p style="color:#1e40af; margin:0 0 8px 0; font-weight:600;">Evaluation adjusted based on project characteristics:</p>'
                                f'<ul style="color:#1e3a5f; margin:0; padding-left:20px; font-size:0.9rem; line-height:1.6;">{reasons_html}</ul>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )

                        if recommendation.get("is_borderline"):
                            _rup_score_badge = (
                                f" &nbsp;<span style='font-size:0.85rem; font-weight:500;'>"
                                f"({int(recommendation.get('recommended_score', 0))} pts vs {int(runner_up_aff)} pts)</span>"
                                if runner_up_aff is not None
                                else ""
                            )
                            st.markdown(
                                f'<div style="background:#fffbeb; border-left:4px solid #f59e0b; border-radius:0 8px 8px 0; padding:16px 20px; margin:0 0 16px 0;">'
                                f'<p style="color:#92400e; margin:0; font-weight:600;">This is a close match between {rec_method} and {runner_up}.{_rup_score_badge}</p>'
                                f'<p style="color:#78350f; margin:6px 0 0 0; font-size:0.9rem;">Review the comparison below to confirm the best fit for this project.</p>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )
                            with st.expander("View Detailed Comparison"):
                                st.markdown(recommendation.get("comparison_text", ""))

                    # === MULTI-METHOD COMPARISON TABLE (HQ only) ===
                    if not is_district_view:
                        _table_label = "All Delivery Methods — Point Ranking" + (
                            (f"  \u00b7  Override-Adjusted ({len(_manual)} change(s) applied)") if _has_overrides else ""
                        )
                        # Build a lookup from original scores for delta comparison
                        _orig_scores = {ms["method"]: ms["score"] for ms in multi_method_data.get("method_scores", [])}
                        with st.expander(_table_label, expanded=True):
                            method_scores = _override_multi.get("method_scores", [])
                            if method_scores:
                                # Helper to convert key factor strings to readable text
                                _sec_names = {
                                    "A": "Project Scope", "B": "Schedule",
                                    "C": "Innovation", "D": "Quality",
                                    "E": "Cost", "F": "Staffing",
                                }
                                def _format_key_factors(kf_list):
                                    parts = []
                                    for kf in kf_list[:4]:
                                        if ":" not in kf:
                                            continue
                                        sec = kf.split(":")[0].strip()
                                        name = _sec_names.get(sec, sec)
                                        if "Strong fit" in kf:
                                            parts.append(f"{name} well-suited")
                                        elif "Poor fit" in kf:
                                            parts.append(f"{name} not well-suited")
                                    return "; ".join(parts) if parts else "Standard criteria apply"

                                mm_rows = []
                                for ms in method_scores:
                                    is_blocked = ms.get("blocked", False)
                                    status = "Ineligible" if is_blocked else "Eligible"

                                    # Key Factors logic:
                                    # Top eligible method -> LLM-generated reasoning
                                    # Ineligible methods  -> show the Caltrans rule that blocked it
                                    # All other eligible  -> human-readable scoring summary
                                    if is_blocked:
                                        block_reasons = ms.get("block_reasons", [])
                                        factors_str = "Blocked by: " + "; ".join(
                                            r.split(": ", 1)[-1] for r in block_reasons
                                        ) if block_reasons else "Does not meet mandatory eligibility criteria"
                                    elif ms.get("key_factors_reasoning"):
                                        factors_str = ms["key_factors_reasoning"]
                                    else:
                                        factors_str = _format_key_factors(ms.get("key_factors", []))

                                    _delta = int(ms['score'] - _orig_scores.get(ms['method'], ms['score']))
                                    mm_rows.append({
                                        "Rank": ms.get("rank", ""),
                                        "Method": ms.get("method", ""),
                                        "Score": int(ms.get('score', 0)),
                                        "Score Change": (
                                            f"+{_delta}" if _delta > 0
                                            else (str(_delta) if _delta != 0 else "—")
                                        ) if _has_overrides else "",
                                        "Status": status,
                                        "Assessment": factors_str,
                                    })
                                mm_df = pd.DataFrame(mm_rows)

                                def _style_method_row(row):
                                    if row.get("Status") == "Ineligible":
                                        return ["color: #9ca3af; font-style: italic;"] * len(row)
                                    rank = row.get("Rank", 99)
                                    if rank == 1:
                                        return ["background-color: #dcfce7; color: #166534; font-weight: bold;"] * len(row)
                                    elif rank <= 3:
                                        return ["background-color: #fef9c3; color: #854d0e;"] * len(row)
                                    return [""] * len(row)

                                # Drop the empty delta column if no overrides
                                if not _has_overrides:
                                    mm_df = mm_df.drop(columns=["Score Change"], errors="ignore")

                                styled_mm = mm_df.style.apply(_style_method_row, axis=1)
                                st.dataframe(styled_mm, use_container_width=True, hide_index=True)

                                # Pros/Cons for top 3 (from override-adjusted data)
                                st.markdown("**Top Methods — Pros & Cons:**")
                                top3 = [ms for ms in _override_multi.get("method_scores", []) if not ms.get("blocked")][:3]
                                for ms in top3:
                                    pros = " • ".join(ms.get("pros", [])[:3])
                                    cons = " • ".join(ms.get("cons", [])[:3])
                                    st.markdown(
                                        f'<div style="border:1px solid #e2e8f0; border-radius:8px; padding:12px 16px; margin-bottom:8px;">'
                                        f'<p style="margin:0 0 4px 0; font-weight:600; color:#1F4E79;">{ms["method"]}</p>'
                                        f'<p style="margin:0; font-size:0.85rem;"><span style="color:#166534;">&#x2705; {pros}</span></p>'
                                        f'<p style="margin:0; font-size:0.85rem;"><span style="color:#991b1b;">&#x26A0; {cons}</span></p>'
                                        f'</div>',
                                        unsafe_allow_html=True,
                                    )

                                bc = multi_method_data.get("borderline_comparison")
                                if bc and bc.get("is_close"):
                                    st.warning(f"⚠ Top methods are within {int(bc['score_gap'])} points of each other — recommend detailed project-specific comparison.")

                    # --- Missing Info Notice (compact — details shown in Step 1 wizard) ---
                    if missing:
                        _sec_map = {"A": "Project Scope", "B": "Schedule", "C": "Innovation", "D": "Quality", "E": "Cost", "F": "Staffing"}
                        _grouped = ", ".join(
                            f"{_sec_map.get(q[0], q[0])} ({sum(1 for x in missing if x[0]==q[0])})"
                            for q in dict.fromkeys(missing)
                        )
                        st.caption(
                            f"📌 **Provisional results** — missing details in: {_grouped}. "
                            "See Step 1 below to add corrections."
                        )

                    # --- Evaluation Rules Reference ---
                    with st.expander("Caltrans Policy Rules Reference", expanded=False):
                        from src.project_delivery_evaluator import OVERRIDE_RULES
                        st.markdown("""
                        <p style="color: #475569; font-size: 0.9rem; margin-bottom: 12px;">
                            The following hardcoded framework rules are applied after scoring to ensure the recommendation
                            aligns with Caltrans delivery method requirements.</p>
                        """, unsafe_allow_html=True)
                        override_status = multi_method_data.get("override_status", [])
                        triggered_ids = [o["rule_id"] for o in override_status if o["triggered"]]
                        for rule in OVERRIDE_RULES:
                            is_trig = rule['id'] in triggered_ids
                            bg_col = "#dcfce7" if is_trig else "#ffffff"
                            bord_col = "#16a34a" if is_trig else "#e2e8f0"
                            lbl = " <span style='color: #166534; font-weight: bold;'>● Applied to this project</span>" if is_trig else ""
                            st.markdown(
                                f'<div style="border:2px solid {bord_col}; background-color:{bg_col}; border-radius:8px; padding:12px 16px; margin-bottom:8px;">'
                                f'<p style="margin:0 0 4px 0; font-weight:600; color:#1F4E79; font-size:0.9rem;">{rule["id"]}: {rule["name"]}{lbl}</p>'
                                f'<p style="margin:0 0 4px 0; color:#64748b; font-size:0.8rem;">Applies when: <code>{rule["trigger"]}</code></p>'
                                f'<p style="margin:0; color:#475569; font-size:0.85rem;">{rule["description"]}</p>'
                                f'</div>',
                                unsafe_allow_html=True,
                            )

                    with st.expander("Institutional Memory (pde_rules.json)", expanded=False):
                        if st.session_state.get("pde_rules"):
                            st.json(st.session_state.pde_rules)
                        else:
                            st.caption("No overrides or learned context found in current rulebook.")

                    if is_district_view:
                        # ============================================================
                        # DISTRICT VIEW — direct Excel download (no HIFL wizard)
                        # ============================================================
                        _dt_excel_mode = st.session_state.get("pde_excel_mode")
                        if "pde_excel_bytes" not in st.session_state or _dt_excel_mode != pde_report_mode:
                            if pde_report_mode == "Template Summary + Method Sheets (V2)":
                                template_path = os.getenv(
                                    "PDE_V2_TEMPLATE_PATH",
                                    "templates/pde_v2_template.xls",
                                )
                                try:
                                    excel_buf = build_evaluation_excel_v2(
                                        eval_result,
                                        recommendation,
                                        project_name,
                                        template_path=template_path,
                                        multi_method_data=multi_method_data,
                                        validation_data=validation_data,
                                        show_points=False,
                                    )
                                    st.session_state.pde_excel_filename = (
                                        f"{project_name.replace(' ', '_')}_delivery_evaluation_v2.xlsx"
                                    )
                                except Exception as _v2err:
                                    st.error(f"Critical error during V2 Excel generation: {_v2err}")
                                    excel_buf = None
                            else:
                                try:
                                    excel_buf = build_evaluation_excel(
                                        eval_result,
                                        recommendation,
                                        project_name,
                                        multi_method_data=multi_method_data,
                                        validation_data=validation_data,
                                    )
                                    st.session_state.pde_excel_filename = (
                                        f"{project_name.replace(' ', '_')}_delivery_evaluation_v1.xlsx"
                                    )
                                except Exception as _v1err:
                                    st.error(f"Critical error during V1 Excel generation: {_v1err}")
                                    excel_buf = None

                            if excel_buf:
                                st.session_state.pde_excel_bytes = excel_buf.getvalue()
                                st.session_state.pde_excel_mode = pde_report_mode
                            else:
                                st.session_state.pde_excel_bytes = None
                                st.session_state.pde_excel_mode = None

                        _dt_dl, _dt_reset, _ = st.columns([1, 1, 3])
                        with _dt_dl:
                            if st.session_state.get("pde_excel_bytes"):
                                st.download_button(
                                    label="⬇ Download Report (.xlsx)",
                                    data=st.session_state.pde_excel_bytes,
                                    file_name=st.session_state.pde_excel_filename,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="pde_download_district",
                                )
                            else:
                                st.warning("Excel report is not available for download.")
                        with _dt_reset:
                            if st.button("🔄 Reset (New Project)", key="pde_reset_district"):
                                for _k in [k for k in list(st.session_state.keys()) if k.startswith("pde_")]:
                                    del st.session_state[_k]
                                st.rerun()

                    if not is_district_view:
                        # ============================================================
                        # HQ WIZARD — 4-step review flow
                        # ============================================================
                        from src.pde_memory_manager import (
                            make_draft_rule,
                            synthesize_rulebook, save_rulebook,
                        )
                        from src.project_delivery_evaluator import generate_key_factors_reasoning

                        # Initialise wizard state
                        if "pde_step" not in st.session_state:
                            st.session_state.pde_step = 1
                        if "pde_manual_ratings" not in st.session_state:
                            st.session_state.pde_manual_ratings = {}
                        if "pde_draft_rules" not in st.session_state:
                            st.session_state.pde_draft_rules = {}
                        if "pde_final_rules" not in st.session_state:
                            st.session_state.pde_final_rules = []
                        if "pde_synthesis_note" not in st.session_state:
                            st.session_state.pde_synthesis_note = ""

                        wizard_step = st.session_state.pde_step

                        # Step indicator bar — active step highlighted, completed steps marked ✓
                        _step_labels = [
                            "1 · Review & Override",
                            "2 · Validation Audit",
                            "3 · Export",
                        ]
                        _si_cols = st.columns(3)
                        for _i, _lbl in enumerate(_step_labels):
                            _step_num = _i + 1
                            _active = (wizard_step == _step_num)
                            _done = (wizard_step > _step_num)
                            _disp_lbl = ("✓ " + _lbl) if _done else _lbl
                            _border_col = "#1F4E79" if _active else ("#16a34a" if _done else "#e2e8f0")
                            _color = "#1F4E79" if _active else ("#15803d" if _done else "#94a3b8")
                            _fw = "700" if _active else "500"
                            _si_cols[_i].markdown(
                                f"<div style='text-align:center; padding:6px 0; "
                                f"border-bottom: 3px solid {_border_col}; "
                                f"color:{_color}; "
                                f"font-weight:{_fw}; "
                                f"font-size:0.85rem;'>{_disp_lbl}</div>",
                                unsafe_allow_html=True,
                            )
                        st.write("")

                        # ----------------------------------------------------------------
                        # STEP 1 — Review ratings, stage corrections
                        # ----------------------------------------------------------------
                        if wizard_step == 1:
                            st.subheader("Step 1 — Review AI Ratings & Provide Corrections")
                            _rating_colors = {
                                "A": ("#dcfce7", "#15803d"),
                                "B": ("#fef9c3", "#854d0e"),
                                "C": ("#fee2e2", "#b91c1c"),
                                "N/E": ("#f1f5f9", "#475569"),
                            }
                            ratings_list = eval_result.get("ratings", [])
                            _missing_qids = [r.get("question_id", "") for r in ratings_list if r.get("missing_info")]

                            # Build per-question rubric lookup so the dropdown shows what A/B/C mean
                            from src.project_delivery_evaluator import RUBRIC_QUESTIONS as _RQ
                            _rubric_lookup = {q["id"]: q for q in _RQ}


                            # --- Missing Info Banner ---
                            if _missing_qids:
                                st.warning(
                                    f"**{len(_missing_qids)} question(s) have missing project information "
                                    f"and need your input:** {', '.join(_missing_qids)}  \n"
                                    f"Select each from the dropdown below to provide the missing details."
                                )

                            # --- Filter toggle ---
                            _filter_mode = st.radio(
                                "Show questions:",
                                ["All Questions", "Missing Info Only"],
                                horizontal=True,
                                key="pde_q_filter",
                            )
                            if _filter_mode == "Missing Info Only":
                                _filtered_ratings = [r for r in ratings_list if r.get("missing_info")]
                                if not _filtered_ratings:
                                    st.success("No missing information detected — all questions have sufficient evidence.")
                                    _filtered_ratings = ratings_list
                            else:
                                _filtered_ratings = ratings_list

                            # Generate dropdown options — show ID + truncated question text for navigation.
                            # The full question text is displayed inside the card below, so we keep the
                            # dropdown concise. Only append "..." when the text was actually truncated.
                            def _q_label(r):
                                flag = " [Missing Info]" if r.get("missing_info") else ""
                                _qt = r.get('question_text', '')
                                _short = (_qt[:52] + "...") if len(_qt) > 52 else _qt
                                return f"{r.get('question_id', '')} — {_short}{flag}"

                            _q_options = [_q_label(r) for r in _filtered_ratings]
                            _sel_q_str = st.selectbox("Select Question:", _q_options, key="pde_q_sel")

                            # Parse QID from the format "A1 — <text>" — split on the em-dash to be safe
                            _sel_qid = _sel_q_str.split(" — ")[0].strip()
                            _r = next((r for r in ratings_list if r.get("question_id", "") == _sel_qid), ratings_list[0])
                        
                            _qid = _r.get("question_id", "")
                            _ai_rating = _r.get("selected_rating", "B").upper()
                            _is_missing = _r.get("missing_info", False)
                            _question_text = _r.get("question_text", "")
                            _source_reasoning = _r.get("source_reasoning", "")
                            _missing_reasoning = _r.get("missing_info_reasoning", "")
                            _confidence = _r.get("confidence", 0)

                            _existing_draft = st.session_state.pde_draft_rules.get(_qid)
                            _current_override = _existing_draft["new_rating"] if _existing_draft else None
                            _display_rating = _current_override if _current_override else _ai_rating
                            _bg, _fg = _rating_colors.get(_display_rating, ("#f1f5f9", "#475569"))

                            with st.container(border=True):
                                _hc1, _hc2 = st.columns([8, 2])
                                with _hc1:
                                    _badge = "  ⚠️ *Missing context*" if _is_missing else ""
                                    st.markdown(f"**{_qid}** — {_question_text}{_badge}")
                                with _hc2:
                                    _override_note = " *(overridden)*" if _current_override else ""
                                    st.markdown(
                                        f"<div style='text-align:center; background:{_bg}; color:{_fg}; "
                                        f"border-radius:6px; padding:4px 0; font-weight:700; font-size:1.1rem;'>"
                                        f"Rating {_display_rating}{_override_note}</div>",
                                        unsafe_allow_html=True,
                                    )

                                # Missing context input
                                if _is_missing:
                                    _mkey = f"pde_missing_{_qid}"
                                    st.text_area(
                                        f"Provide missing context for {_qid}:",
                                        value=st.session_state.get(_mkey, ""),
                                        placeholder=_missing_reasoning or "Describe what is missing...",
                                        key=_mkey,
                                        height=70,
                                    )

                                # Evidence (collapsed by default — user can expand if needed)
                                with st.expander(
                                    f"View AI Evidence (Confidence: {_confidence:.0%})",
                                    expanded=False,
                                ):
                                    st.caption(_source_reasoning)
                                    if _missing_reasoning and "None" not in _missing_reasoning:
                                        st.caption(f"**If resolved:** {_missing_reasoning}")

                                _rubric_q = _rubric_lookup.get(_qid, {})
                                _option_labels = {
                                    "A": f"A — {_rubric_q['option_a']}" if _rubric_q.get('option_a') else "A",
                                    "B": f"B — {_rubric_q['option_b']}" if _rubric_q.get('option_b') else "B",
                                    "C": f"C — {_rubric_q['option_c']}" if _rubric_q.get('option_c') else "C",
                                }
                                _default_rating = _current_override or _ai_rating
                                _default_idx = ["A", "B", "C"].index(_default_rating) if _default_rating in ("A", "B", "C") else 0
                                _nr = st.selectbox(
                                    "Your Rating Override",
                                    ["A", "B", "C"],
                                    index=_default_idx,
                                    format_func=lambda x, _lbl=_option_labels: _lbl[x],
                                    key=f"pde_or_{_qid}",
                                )
                                _nreason = st.text_area(
                                    "Why are you changing this? (required to save)",
                                    value=_existing_draft.get("user_rationale", "") if _existing_draft else "",
                                    placeholder="e.g. 'Section 4 confirms 70% design complete, not 30% as assumed.'",
                                    key=f"pde_reason_{_qid}",
                                    height=70,
                                )
                            
                                if st.button("Save Override", key=f"pde_save_{_qid}", type="primary"):
                                    if len(_nreason.strip()) < 15:
                                        st.warning("Please provide a more specific reason (at least 15 characters).")
                                    elif _nr == _ai_rating:
                                        st.warning("Rating unchanged — select a different rating to save an override.")
                                    else:
                                        _draft = make_draft_rule(
                                            question_id=_qid,
                                            summary=f"Rating changed from {_ai_rating} to {_nr}",
                                            source_evidence=_source_reasoning[:400] or "Not available",
                                            user_rationale=_nreason.strip(),
                                        )
                                        _draft["new_rating"] = _nr
                                        _draft["original_rating"] = _ai_rating
                                        st.session_state.pde_draft_rules[_qid] = _draft
                                        st.session_state.pde_manual_ratings[_qid] = _nr
                                        st.session_state.pde_existing_ratings = st.session_state.pde_manual_ratings
                                        for _stale in ["pde_validation", "pde_multi_method", "pde_override_multi", "pde_override_rec", "pde_override_multi_key"]:
                                            if _stale in st.session_state:
                                                del st.session_state[_stale]
                                        st.rerun()

                            st.write("---")
                            st.subheader("Pending Corrections")
                            if not st.session_state.pde_draft_rules:
                                st.caption("No overrides staged yet. You may proceed to Validation.")
                            else:
                                for _dqid, _drule in list(st.session_state.pde_draft_rules.items()):
                                    with st.container(border=True):
                                        _qc1, _qc2 = st.columns([8, 2])
                                        with _qc1:
                                            st.markdown(f"**{_dqid}** changed from `{_drule['original_rating']}` to `{_drule['new_rating']}`")
                                            st.caption(f"*Reason:* {_drule.get('user_rationale', '')}")
                                        with _qc2:
                                            if st.button("✕ Remove", key=f"pde_rm_{_dqid}"):
                                                del st.session_state.pde_draft_rules[_dqid]
                                                del st.session_state.pde_manual_ratings[_dqid]
                                                st.session_state.pde_existing_ratings = st.session_state.pde_manual_ratings
                                                for _stale in ["pde_validation", "pde_multi_method", "pde_override_multi", "pde_override_rec", "pde_override_multi_key"]:
                                                    if _stale in st.session_state:
                                                        del st.session_state[_stale]
                                                st.rerun()

                            st.write("")
                            if st.session_state.pde_draft_rules:
                                st.info(
                                    f"{len(st.session_state.pde_draft_rules)} correction(s) staged. "
                                    "Proceed to Validation Audit to cross-check before export."
                                )
                            else:
                                st.info("No corrections staged — you may proceed directly to export.")
                            _n1, _n2, _n3 = st.columns([1, 1, 3])
                            with _n1:
                                if st.button("Next → Validation", key="pde_to2", type="primary"):
                                    st.session_state.pde_step = 2
                                    st.rerun()

                        # ----------------------------------------------------------------
                        # STEP 2 — Validation Audit (mandatory)
                        # ----------------------------------------------------------------
                        elif wizard_step == 2:
                            st.subheader("Step 2 — Validation Audit")
                            st.caption(
                                "Review the AI's assessment of each correction you have staged. "
                                "This step is mandatory before export to ensure each rating change is defensible "
                                "and transparently logged into institutional memory."
                            )

                            if "pde_validation" not in st.session_state:
                                if st.button("Run Audit", type="primary", key="pde_run_val"):
                                    with st.spinner("Running validation analysis..."):
                                        from src.project_delivery_evaluator import run_validation_analysis
                                        _val = run_validation_analysis(
                                            eval_result.get("ratings", []),
                                            st.session_state.pde_manual_ratings,
                                        )
                                    st.session_state.pde_validation = _val
                                    st.rerun()
                            else:
                                _val = st.session_state.pde_validation
                                _vs = _val.get("summary", {})
                                _vi = _val.get("deviation_impact", {})

                                # Summary metrics row
                                _rate = _vs.get('agreement_rate', 0)
                                _rate_color = "#166534" if _rate >= 80 else ("#b45309" if _rate >= 50 else "#991b1b")
                                _vm1, _vm2, _vm3, _vm4 = st.columns(4)
                                _vm1.metric("Agreement Rate", f"{_rate:.1f}%")
                                _vm2.metric("Matches", _vs.get("matches", 0))
                                _vm3.metric("Minor Disagreements", _vs.get("minor_mismatches", 0))
                                _vm4.metric("Major Disagreements", _vs.get("major_mismatches", 0))

                                if _vi.get("recommendation_changed"):
                                    st.warning(
                                        f"Your overrides shift the final recommendation: "
                                        f"**{_vi['ai_method']}** (AI) → **{_vi['user_method']}** (Your Ratings)"
                                    )
                                else:
                                    st.success(f"Final recommendation stays **{_vi.get('ai_method', '')}** regardless of overrides.")

                                if _rate < 10:
                                    st.warning(
                                        "Agreement rate is below 10%. A high number of ratings differ from the AI's "
                                        "initial assessment. Please provide a brief project-level context note below "
                                        "to document the reason for these changes."
                                    )
                                    st.text_area(
                                        "Global Context (explain the paradigm shift):",
                                        key="pde_global_context",
                                        placeholder="e.g. 'This project uses a non-standard fast-track procurement, which changes the applicability of several standard criteria.'",
                                        height=90
                                    )

                                # Per-correction audit cards
                                _mismatches = _val.get("mismatches", [])
                                _all_comps = {c["question_id"]: c for c in _val.get("comparisons", [])}
                                _staged = st.session_state.pde_draft_rules

                                if _staged:
                                    st.markdown("---")
                                    st.markdown("**Detailed Correction Review**")
                                    st.caption("For each correction you staged, the AI's original reasoning and confidence are shown alongside your rationale.")

                                    for _dqid, _drule in _staged.items():
                                        _comp = _all_comps.get(_dqid, {})
                                        _sev = _comp.get("severity", "match")
                                        _conf = _comp.get("ai_confidence", 0)
                                        _ai_ev = _comp.get("ai_evidence", "No AI reasoning available.")
                                        _ur = _drule.get("user_rationale", "No rationale recorded.")
                                        _ai_r = _drule.get("original_rating", "?")
                                        _u_r = _drule.get("new_rating", "?")

                                        if _sev == "major_mismatch":
                                            _conf_val = _conf
                                            if _conf_val >= 0.75:
                                                _sev_color = "#991b1b"
                                                _sev_label = "High-Confidence Override"
                                            else:
                                                _sev_color = "#c2410c"
                                                _sev_label = "Major Difference"
                                            _sev_bg = "#fef2f2"
                                            _show_reasoning = True
                                        elif _sev == "minor_mismatch":
                                            _sev_color = "#b45309"
                                            _sev_label = "Minor Difference"
                                            _sev_bg = "#fffbeb"
                                            _show_reasoning = True
                                        else:
                                            _sev_color = "#166534"
                                            _sev_label = "AI Agrees"
                                            _sev_bg = "#f0fdf4"
                                            _show_reasoning = False

                                        # Neutral AI reasoning block — shows the AI's original evidence as-is.
                                        # NOT a rebuttal; we do not try to argue against the user's rationale
                                        # because the AI evidence can accidentally support the user's own argument.
                                        if _show_reasoning:
                                            _ai_ev_short = (_ai_ev[:350] + "…") if len(_ai_ev) > 350 else _ai_ev
                                            _counter_html = (
                                                f'<div style="background:#f8fafc;border-radius:6px;padding:10px;margin-bottom:8px;border-left:3px solid {_sev_color};">'
                                                f'<p style="margin:0 0 5px 0;font-size:0.75rem;text-transform:uppercase;color:{_sev_color};font-weight:600;">AI\'s Original Reasoning (rated {_ai_r}):</p>'
                                                f'<p style="margin:0;font-size:0.85rem;color:#374151;">{_ai_ev_short}</p>'
                                                f'</div>'
                                            )
                                        else:
                                            _counter_html = ""


                                        st.markdown(
                                            f'<div style="border:1px solid {_sev_color}40; border-left:4px solid {_sev_color}; background:{_sev_bg}; border-radius:8px; padding:14px 18px; margin-bottom:12px;">'
                                            f'<div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">'
                                            f'<span style="font-weight:700; color:#1F4E79; font-size:1rem;">{_dqid}</span>'
                                            f'<span style="font-size:0.8rem; color:{_sev_color}; font-weight:600;">{_sev_label}</span>'
                                            f'</div>'
                                            f'<div style="display:flex; gap:12px; margin-bottom:10px; font-size:0.85rem;">'
                                            f'<span>AI Rating: <strong>{_ai_r}</strong></span>'
                                            f'<span>&rarr;</span>'
                                            f'<span>Your Rating: <strong>{_u_r}</strong></span>'
                                            f'<span style="margin-left:auto; color:#64748b;">AI Confidence: {_conf:.0%}</span>'
                                            f'</div>'
                                            f'{_counter_html}'
                                            f'<div style="background:#eff6ff; border-radius:6px; padding:10px;">'
                                            f'<p style="margin:0 0 4px 0; font-size:0.75rem; text-transform:uppercase; color:#1e40af; font-weight:600;">Your Override Rationale</p>'
                                            f'<p style="margin:0; font-size:0.85rem; color:#1e3a8a;">{_ur}</p>'
                                            f'</div>'
                                            f'</div>',
                                            unsafe_allow_html=True,
                                        )
                                else:
                                    st.info("No corrections staged — all AI ratings accepted as-is.")

                            _v1, _v2, _v3 = st.columns([1, 1, 3])
                            with _v1:
                                if st.button("← Back", key="pde_vback"):
                                    # Clear stale validation so Step 2 always shows fresh on return
                                    for _stale in ["pde_validation"]:
                                        if _stale in st.session_state:
                                            del st.session_state[_stale]
                                    st.session_state.pde_step = 1
                                    st.rerun()
                            with _v2:
                                if st.button("Finalize & Export →", type="primary", key="pde_v2to3"):
                                    # If global context was added, maybe package it as a master rule or simply ignore its logic implementation for prototype.
                                    # Since we dropping security, we auto-approve everything here too
                                    _approved_rules = list(st.session_state.pde_draft_rules.values())
                                
                                    # Package global context if they wrote it
                                    if st.session_state.get("pde_global_context"):
                                        _ctx_rule = make_draft_rule(
                                            question_id="GLOBAL",
                                            summary="Agreement rate below 10% — override context provided",
                                            source_evidence="Agreement dropped below 10%",
                                            user_rationale=st.session_state.pde_global_context
                                        )
                                        _ctx_rule["status"] = "approved"
                                        _approved_rules.append(_ctx_rule)

                                    for _r in _approved_rules:
                                        _r["status"] = "approved"
                                    st.session_state.pde_final_rules = _approved_rules

                                    _existing_rb = st.session_state.get("pde_rules", [])
                                    if _approved_rules:
                                        with st.spinner("Synthesizing rulebook..."):
                                            _merged, _note = synthesize_rulebook(_existing_rb, _approved_rules)
                                        st.session_state.pde_rules = _merged
                                        st.session_state.pde_synthesis_note = _note
                                        for _stale in ["pde_excel_bytes", "pde_excel_mode"]:
                                            if _stale in st.session_state:
                                                del st.session_state[_stale]
                                    st.session_state.pde_step = 3
                                    st.rerun()

                        # ----------------------------------------------------------------
                        # STEP 3 — Export (Excel + Rulebook)
                        # ----------------------------------------------------------------
                        elif wizard_step == 3:
                            st.subheader("Step 3 — Export Report & Rulebook")

                            _synth_note = st.session_state.get("pde_synthesis_note", "")
                            if _synth_note:
                                st.info(f"📚 {_synth_note}")

                            _fc4 = len(st.session_state.pde_final_rules)
                            if _fc4:
                                st.success(
                                    f"**{_fc4} correction(s) committed to institutional memory.** "
                                    "These have been merged into the project rulebook."
                                )
                                with st.expander("View committed corrections", expanded=False):
                                    for _r in st.session_state.pde_final_rules:
                                        _orig = _r.get('original_rating', '')
                                        _new = _r.get('new_rating', '')
                                        _qid_label = _r.get('question_id', '')
                                        _rationale_short = _r.get('user_rationale', '')[:100]
                                        if _orig and _new:
                                            _change_label = f"Rating changed from {_orig} to {_new}"
                                        else:
                                            _change_label = _r.get('summary', '')
                                        st.markdown(
                                            f"- **{_qid_label}** — {_change_label}" +
                                            (f" · *{_rationale_short}{'...' if len(_r.get('user_rationale','')) > 100 else ''}*" if _rationale_short else "")
                                        )

                            # Excel generation (cached)
                            excel_cache_mode = st.session_state.get("pde_excel_mode")
                            if "pde_excel_bytes" not in st.session_state or excel_cache_mode != pde_report_mode:
                                if pde_report_mode == "Template Summary + Method Sheets (V2)":
                                    template_path = os.getenv(
                                        "PDE_V2_TEMPLATE_PATH",
                                        "templates/pde_v2_template.xls",
                                    )
                                    try:
                                        # Build a patched eval_result that carries override-adjusted ratings
                                        _export_eval = dict(eval_result)
                                        _export_eval["ratings"] = _merged_ratings
                                        excel_buf = build_evaluation_excel_v2(
                                            _export_eval,
                                            _override_rec,
                                            project_name,
                                            template_path=template_path,
                                            multi_method_data=_override_multi,
                                            validation_data=validation_data,
                                        )
                                        st.session_state.pde_excel_filename = (
                                            f"{project_name.replace(' ', '_')}_delivery_evaluation_v2.xlsx"
                                        )
                                    except Exception as _v2err:
                                        st.error(f"Critical error during V2 Excel generation: {_v2err}")
                                        excel_buf = None
                                else:
                                    try:
                                        # Build a patched eval_result that carries override-adjusted ratings
                                        _export_eval = dict(eval_result)
                                        _export_eval["ratings"] = _merged_ratings
                                        excel_buf = build_evaluation_excel(
                                            _export_eval,
                                            _override_rec,
                                            project_name,
                                            multi_method_data=_override_multi,
                                            validation_data=validation_data,
                                        )
                                        st.session_state.pde_excel_filename = (
                                            f"{project_name.replace(' ', '_')}_delivery_evaluation_v1.xlsx"
                                        )
                                    except Exception as _v1err:
                                        st.error(f"Critical error during V1 Excel generation: {_v1err}")
                                        excel_buf = None

                                if excel_buf:
                                    st.session_state.pde_excel_bytes = excel_buf.getvalue()
                                    st.session_state.pde_excel_mode = pde_report_mode
                                else:
                                    st.session_state.pde_excel_bytes = None
                                    st.session_state.pde_excel_mode = None

                            # Download buttons
                            _dl1, _dl2, _dl3 = st.columns([1, 1, 2])
                            with _dl1:
                                if st.session_state.get("pde_excel_bytes"):
                                    st.download_button(
                                        label="⬇ Download Report (.xlsx)",
                                        data=st.session_state.pde_excel_bytes,
                                        file_name=st.session_state.pde_excel_filename,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key="pde_download_excel",
                                    )
                                else:
                                    st.warning("Excel report not available.")
                            with _dl2:
                                _all_rules = st.session_state.get("pde_rules", [])
                                _approved_rules = [r for r in _all_rules if r.get("status") == "approved"]
                                if _approved_rules:
                                    _rb_json = save_rulebook(_all_rules)
                                    st.download_button(
                                        label="⬇ Download Rulebook (.json)",
                                        data=_rb_json,
                                        file_name="pde_rules.json",
                                        mime="application/json",
                                        key="pde_download_rulebook",
                                    )
                                else:
                                    st.caption("No approved rules to export yet.")

                            # Reset
                            st.write("")
                            if st.session_state.pde_final_rules:
                                st.caption(
                                    "To add further corrections, use the Back button. "
                                    "Click Finalize & Export again to update the report."
                                )
                            if st.button("← Back to Review", key="pde_bk4"):
                                # Clear Excel cache so re-entry regenerates a fresh report
                                for _stale in ["pde_excel_bytes", "pde_excel_mode"]:
                                    if _stale in st.session_state:
                                        del st.session_state[_stale]
                                st.session_state.pde_step = 1
                                st.rerun()
                            if st.button("🔄 Reset (New Project)", key="pde_reset"):
                                for _k in [k for k in list(st.session_state.keys()) if k.startswith("pde_")]:
                                    del st.session_state[_k]
                                st.rerun()
        with col73:
            st.write("")

    else:
        with col22:
            st.subheader("Scope of Data Exchange")
            st.write("#")
        with col24:
            select_all = st.checkbox("Select All")

        base_fields_list = [
            "Actual release date",
            "Name of the youth",
            "Race/Ethnicity",
            "Medi-Cal ID Number",
            "Residential Address",
            "Telephone",
            "Medi-Cal health plan assigned",
            "Health Screenings",
            "Health Assessments",
            "Chronic Conditions",
            "Prescribed Medications",
        ]
